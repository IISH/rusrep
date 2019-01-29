#!/usr/bin/python
# -*- coding: utf-8 -*-

# VT-07-Jul-2016 Latest change by VT
# FL-18-Apr-2018 handle None cursor
# FL-24-Apr-2018 GridFS
# FL-29-Jan-2019 aggregate_dataset: fields (old) & records (new) versions

import gridfs
import json
import logging
import openpyxl
import os
import re
import sys

from collections import OrderedDict
from datetime import date
from icu import Locale, Collator
from operator import itemgetter
import pandas as pd
from pymongo import MongoClient
from sys import exc_info

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def preprocessor( use_gridfs, datafilter ):
    logging.info( "preprocessor() datafilter: %s" % datafilter )
    
    lex_lands        = {}
    vocab_regs_terms = {}           # regions and terms
    sheet_header     = []           # top rows of download spreadsheet
    topic_name       = ""
    params           = {}
    
    key = datafilter.get( "key" )
    if not key:
        logging.error( "preprocessor() NO key?" )
        return ( lex_lands, vocab_regs_terms, sheet_header, topic_name, params )
    
    key_comps = key.split( '-' )
    logging.debug( "key_comps: %s" % str( key_comps ) )
    
    lang     = ''       # { 'en' | 'ru' }
    hist_mod = ''       # { 'h' | 'm' }
    topic    = ''       # x.yz
    
    if len( key_comps ) >= 1:
        lang = key_comps[ 0 ]
    if len( key_comps ) >= 2:
        hist_mod = key_comps[ 1 ]
    if len( key_comps ) >= 3:
        topic = key_comps[ 2 ]
    
    clientcache = MongoClient()
    db_cache = clientcache.get_database( 'datacache' )
    logging.debug( "find cached data with key: %s" % key )
    
    cache_data = {}
    if use_gridfs:
        fs_cache = gridfs.GridFS( db_cache )
        result_str = fs_cache.get( str( key ) ).read()
        cache_data = json.loads( result_str )
    else:
        cursor = db_cache.data.find_one( { "key": key } )
        if cursor:
            cache_data = dict ( cursor )
        else:
            logging.error( "preprocessor() no cursor data for key: %s" % key )
    
    logging.debug( "# keys in cache_data: %s" % len( cache_data.keys() ) )
    for key, value in cache_data.iteritems():
        if isinstance( value, list ):
            logging.debug( "key %s: value type: %s, # of elements: %d" % ( key, type( value ), len( value ) ) )
        elif isinstance( value, dict ):
            logging.debug( "key %s: value type: %s, # of keys: %d" % ( key, type( value ), len( value ) ) )
        else:
            logging.debug( "key %s: value type: %s" % ( key, type( value ) ) )
    
    nkeys = 0
    params = cache_data.get( "params" )
    if params:
        nkeys = len( params.keys() )
        logging.debug( "# of keys in params: %d" % nkeys )
        for key, value in params.iteritems():
            logging.debug( "key %s: value: %s" % ( key, value ) )
    else:
        logging.error( "No params in data cache" )
    
    data = cache_data.get( "data" )
    
    base_year = '0'
    ter_codes = []      # actually used region codes
    dataset   = []
    
    if data:
        for i, item in enumerate( data ):  
            logging.debug( "%d, item type: %s, %s" % ( i, type( item ), str( item ) ) )
            
            dataitem = item
            if 'path' in item:
                classes = item[ 'path' ]
                del item[ 'path' ]
                clist = {}
                for classname in classes:
                    dataitem[ classname] = classes[ classname ]
            if 'base_year' in item:
                base_year = item[ 'base_year' ]
            
            item_lexicon = dataitem
            lands = {}
            if 'ter_code' in item_lexicon:
                ter_code = item_lexicon[ 'ter_code' ]
                ter_codes.append( ter_code )
                lands[ item_lexicon[ 'ter_code' ] ] = item_lexicon.get( "total" )
                del item_lexicon[ 'ter_code' ]
            if 'total' in item_lexicon:
                del item_lexicon[ 'total' ]
            
            try:
                count = item_lexicon[ 'count' ]
                del item_lexicon[ 'count' ]          # 'count' should not be part of lex_key
            except:
                pass
            
            # lex_key as stringyfied ordered dict
            keys = sorted( item_lexicon.iterkeys() )                    # sorted keys
            tuples = [ ( key, item_lexicon[ key ] ) for key in keys ]   # tuple list ordered by keys
            ordered_dict = OrderedDict( tuples )                        # dict ordered by keys
            lex_key = json.dumps( ordered_dict )                        # lex_key as string
            
            item_lexicon[ 'lands' ] = lands
            if lex_key in lex_lands:
                logging.debug( "old lex_key: %s" % lex_key )
                current_lands = lex_lands[ lex_key ]
                for item in lands:
                    current_lands[ item ] = lands[ item ]
                lex_lands[ lex_key ] = current_lands
            else:
                logging.debug( "new lex_key: %s" % lex_key )
                lex_lands[ lex_key ] = lands
            
            dataset.append( dataitem )
        
        if len( ter_codes ) == 0:
            ter_codes = params.get( "ter_codes" )
        vocab_regs_terms[ "ter_codes" ] = ter_codes
        
        db_vocabulary = clientcache.get_database( 'vocabulary' )   # vocabulary
        
        # create sheet_header
        # load terms
        terms_needed  = [ "na", "base_year", "count", "datatype", "value_unit" ]
        if hist_mod == 'h':     # historical
            terms_needed += [ "histclass1", "histclass2", "histclass3", "histclass4", "histclass5", "histclass6", "histclass7", "histclass8", "histclass9", "histclass10" ]
        elif hist_mod == 'm':   # modern
            terms_needed += [ "class1", "class2", "class3", "class4", "class5", "class6", "class7", "class8", "class9", "class10" ]
        
        terms = {}
        vocab_download = db_vocabulary.data.find( { "vocabulary": "ERRHS_Vocabulary_download" } )
        
        for item in vocab_download:
            #logging.debug( str( item ) )
            ID = item[ 'ID' ]
            #logging.debug( "%s %s" % ( topic, ID ) )
            if ID == topic:
                if lang == 'en':
                    topic_name = item[ 'EN' ]
                else:
                    topic_name = item[ 'RUS' ]
            
            if ID in terms_needed:
                if lang == 'en':
                    terms[ item[ 'ID' ] ] = item[ 'EN' ]
                else:
                    terms[ item[ 'ID' ] ] = item[ 'RUS' ]
        
        #logging.debug( "topic: %s, topic_name: %s" % ( topic, topic_name ) )
        vocab_regs_terms[ "terms" ] = terms
        
    if lang == 'en':
        if hist_mod == 'h':
            classification = "HISTORICAL"
        elif hist_mod == 'm':
            classification = "MODERN"
        else:
            classification = ""
        
        sheet_header.append( { "r" : 1, "c" : 0, "value" : "Electronic repository of Russian Historical Statistics - ristat.org" } )
        sheet_header.append( { "r" : 3, "c" : 0, "value" : "TOPIC:" } )
        sheet_header.append( { "r" : 3, "c" : 1, "value" : topic_name } )
        sheet_header.append( { "r" : 4, "c" : 0, "value" : "BENCHMARK-YEAR:" } )
        sheet_header.append( { "r" : 4, "c" : 1, "value" : base_year } )
        sheet_header.append( { "r" : 5, "c" : 0, "value" : "CLASSIFICATION:" } )
        sheet_header.append( { "r" : 5, "c" : 1, "value" : classification } )
        sheet_header.append( { "r" : 6, "c" : 0, "value" : "NUMBER" } )
        sheet_header.append( { "r" : 6, "c" : 1, "value" : 0 } )
    else:
        if hist_mod == 'h':
            classification = "ИСТОРИЧЕСКАЯ"
        elif hist_mod == 'm':
            classification = "СОВРЕМЕННАЯ"
        else:
            classification = ""
        
        sheet_header.append( { "r" : 1, "c" : 0, "value" : "Электронный архив Российской исторической статистики - ristat.org" } )
        sheet_header.append( { "r" : 3, "c" : 0, "value" : "ТЕМА:" } )
        sheet_header.append( { "r" : 3, "c" : 1, "value" : topic_name } )
        sheet_header.append( { "r" : 4, "c" : 0, "value" : "ГОД:" } )
        sheet_header.append( { "r" : 4, "c" : 1, "value" : base_year } )
        sheet_header.append( { "r" : 5, "c" : 0, "value" : "КЛАССИФИКАЦИЯ:" } )
        sheet_header.append( { "r" : 5, "c" : 1, "value" : classification } )
        sheet_header.append( { "r" : 6, "c" : 0, "value" : "ЧИСЛО" } )
        sheet_header.append( { "r" : 6, "c" : 1, "value" : 0 } )
        
    if not cache_data:
        # this happens when the data could not be cached in MongoDB because its size exceeded the limit
        sheet_header.append( { "r" : 7, "c" : 0, "value" : "Zero items received from MongoDB for key:" } )
        sheet_header.append( { "r" : 7, "c" : 1, "value" : key } )
        sheet_header.append( { "r" : 8, "c" : 0, "value" : "MongoDB CACHING FAILED?, BSON document SIZE TOO LARGE?" } )
    
    logging.debug( "preprocessor (%d) dataset: %s"          % ( len( dataset ),          str( dataset ) ) )
    logging.debug( "preprocessor (%d) ter_codes: %s"        % ( len( ter_codes ),        str( ter_codes ) ) )
    logging.debug( "preprocessor (%d) lex_lands: %s"        % ( len( lex_lands ),        str( lex_lands ) ) )
    logging.debug( "preprocessor (%d) vocab_regs_terms: %s" % ( len( vocab_regs_terms ), str( vocab_regs_terms ) ) )
    logging.debug( "preprocessor (%d) sheet_header: %s"     % ( len( sheet_header ),     str( sheet_header ) ) )
    
    return lex_lands, vocab_regs_terms, sheet_header, topic_name, params



def aggregate_dataset_fields( key, download_dir, xlsx_name, lex_lands, vocab_regs_terms, sheet_header, topic_name, params ):
    logging.info( "aggregate_dataset_fields()" )
    logging.debug( "key: %s" % key )
    logging.debug( "download_dir: %s" % download_dir )
    logging.debug( "xlsx_name: %s" % xlsx_name )
    logging.debug( "lex_lands: %s" % str( lex_lands ) )
    logging.debug( "vocab_regs_terms: %s" % str( vocab_regs_terms ) )
    logging.debug( "sheet_header type: %s" % type( sheet_header ) )
    logging.debug( "topic_name: %s" % topic_name )
    logging.debug( "params: %s" % str( params ) )
    
    xlsx_pathname = ""
    if not os.path.isdir( download_dir ):
        msg = "download destination was removed"
        return xlsx_pathname, msg

    xlsx_pathname = os.path.abspath( os.path.join( download_dir, xlsx_name ) )
    logging.debug( "full_path: %s" % xlsx_pathname )
    
    lang = ""
    hist_mod = ""
    base_year_str = ""
    key_comps = key.split( '-' )
    if len( key_comps ) >= 1:
        lang = key_comps[ 0 ]
    if len( key_comps ) >= 2:
        hist_mod = key_comps[ 1 ]
    if len( key_comps ) >= 3:
        base_year_str = key_comps[ 3 ]
        base_year = int( base_year_str )
    logging.debug( "lang: %s, hist_mod: %s, base_year_str: %s" % ( lang, hist_mod, base_year_str ) )
    
    logging.debug( "# of keys in vocab_regs_terms: %d" % len( vocab_regs_terms.keys() ) )
    for k, v in vocab_regs_terms.iteritems():
        logging.debug( "k: %s, v: %s" % ( k, str( v ) ) )
    
    nsheets = 0
    ter_code_list = []
    
    #wb = openpyxl.Workbook( encoding = "utf-8" )
    wb = openpyxl.Workbook()
    
    escape = False
    if escape:
        logging.debug( "openpyxl version: %s" % openpyxl.__version__ )
        logging.debug( "empty spreadsheet" )
        ws = wb.active                  # grab the active worksheet
        ws['A1'] = 42                   # Data can be assigned directly to cells
        ws.append([1, 2, 3])            # Rows can also be appended
        import datetime                 # Python types will automatically be converted
        ws['A2'] = datetime.datetime.now()
        wb.save( "/home/dpe/tmp/data/download/sample.xlsx" )
        wb.save( xlsx_pathname )    # Save the file
        return xlsx_pathname, ""
    
    clientcache = MongoClient()
    db_cache = clientcache.get_database( 'datacache' )
    
    #logging.debug( "find cached data with key: %s" % key )
    #result = db_cache.data.find( { "key": key } )
    #logging.info( "aggregate_dataset_fields() length of cached dict: %d for key: %s" % ( len( result ), key ) )
    
    db_vocabulary = clientcache.get_database( 'vocabulary' )   # vocabulary
    
    logging.debug( "params: %s, %s" % ( params, type( params ) ) )
    if params:
        for k in params:
            logging.debug( "key: %s, value: %s" % ( k, params[ k ] ) )
    
        try:
            ter_code_list = params[ "ter_codes" ]
        except:
            ter_code_list = []
    
        logging.debug( "ter_code_list: %s, %s" % ( str( ter_code_list ), type( ter_code_list ) ) )
        
        try:
            level_paths = params[ "path" ]
        except:
            level_paths = []
        
        logging.debug( "# of level paths: %d" % len( level_paths ) )
        for level_path in level_paths:
            logging.debug( "level path: %s" % level_path )
    else:
        logging.error( "aggregate_dataset_fields() NO params" )
    
    nrecords = []       # number of data records per sheet
    if hist_mod == 'h':
        nsheets = 1
        base_years = [ base_year ]
        ws_0 = wb.get_active_sheet()
        ws_0.title = "Dataset"

    elif hist_mod == 'm':
        nsheets = 5
        base_years = [ "1795", "1858", "1897", "1959", "2002" ]
        ws_0 = wb.get_active_sheet()
        ws_0.title = "1795"
        ws_1 = wb.create_sheet( "1858", 1 )
        ws_2 = wb.create_sheet( "1897", 2 )
        ws_3 = wb.create_sheet( "1959", 3 )
        ws_4 = wb.create_sheet( "2002", 4 )
        
        vocab_regions = db_vocabulary.data.find( { "vocabulary": "ERRHS_Vocabulary_regions" } )
        #logging.debug( "vocab_regions: %s" % vocab_regions )
        ter_code_list = []
        for region in vocab_regions:
            logging.debug( "region: %s" % str( region ) )
            ter_code_list.append( region[ "ID" ] )
    
    # initialise base_year counts
    byear_counts = { "1795" : 0, "1858" : 0, "1897" : 0, "1959" : 0, "2002" : 0 }
    
    #skip_list = [ "count" ]
    skip_list = []
    
    # loop over the data sheets, and select the correct year data
    for sheet_idx in range( nsheets ):
        base_year = base_years[ sheet_idx ]
        logging.debug( "work_sheet: %d, base_year: %s" % ( sheet_idx, base_year ) )
        if sheet_idx == 0:
            ws = ws_0
        elif sheet_idx == 1:
            ws = ws_1
        elif sheet_idx == 2:
            ws = ws_2
        elif sheet_idx == 3:
            ws = ws_3
        elif sheet_idx == 4:
            ws = ws_4
        
        # load regions for base_year
        regions_filter = {}
        regions_filter[ "vocabulary" ] = "ERRHS_Vocabulary_regions"
        # regions depend on base_year
        regions_filter[ 'basisyear' ] = str( base_year )    # basisyear is int column in vocabulary file
        logging.debug( "regions_filter: %s" % str( regions_filter ) )
        cursor = db_vocabulary.data.find( regions_filter )
        logging.debug( "# of regions from vocabulary: %d" % cursor.count() )
        vocab_regions = list( cursor )
        
        # get region names from ter_codes
        regions = {}
        #ter_codes = vocab_regs_terms[ "ter_codes" ]    # from sql result
        ter_codes = ter_code_list                       # from qinput or all (modern)
        logging.debug( "ter_codes: %s" % str( ter_codes ) )
        for i, item in enumerate( vocab_regions ):
            logging.debug( "%d: %s" % ( i, str( item ) ) )
            ID = item[ 'ID' ]
            if ID in ter_codes:
                if lang == 'en':
                    regions[ item[ 'ID' ] ] = item[ 'EN' ]
                else:
                    regions[ item[ 'ID' ] ] = item[ 'RUS' ]
        
        logging.debug( "regions: %s" % str( regions ) )
        vocab_regs_terms[ "regions" ] = regions
        
        logging.debug( "%d keys in vocab_regs_terms:" % len( vocab_regs_terms ) )
        for key in vocab_regs_terms:
            logging.debug( "key: %s" % key )
        
        try:
            na = vocab_regs_terms[ "terms" ][ "na" ]
        except:
            na = "na"
        logging.debug( "na: %s" % na )
        
        regions = {}        # key: reg_name, value: reg_code
        reg_names = []      # list of reg_name's
        for ter_code in vocab_regs_terms[ "regions" ]:
            tmpname = vocab_regs_terms[ "regions" ][ ter_code ]
            ter_name = tmpname.decode( 'utf-8' )
            regions[ ter_name ] = ter_code
            reg_names.append( ter_name ) 
            
        logging.debug( "unsorted reg_names: %s" % str( reg_names ) )
        # sort the region names alphabetically
        locale = Locale( 'ru' )  # 'el' is the locale code for Greek
        collator = Collator.createInstance( locale )
        sorted_regions = sorted( reg_names, cmp = collator.compare )
        logging.debug( "sorted_regions: %s" % str( sorted_regions ) )
        
        # empty level strings are not in the data, but we need same # of colums 
        # for all rows; find the maximum number of levels
        header_chain = set()
        class_chain  = set()
        lex_lands_list0 = []
        for itemchain in lex_lands:
            chain = json.loads( itemchain )
            lex_lands_list0.append( chain )
            
            for name in sorted( chain ):
                header_chain.add( name )
                if "class" in name:
                    class_chain.add( name )
        
        lex_lands_list = []
        for chain in lex_lands_list0:
            for class_ in class_chain:
                if not chain.get( class_ ):
                    chain[ class_ ] = ''
            lex_lands_list.append( chain )
            logging.debug( "chain: %s" % str( chain ) )
        
        #max_columns = len( header_chain )
        #max_columns = 1 + len( header_chain )       # +1: for count column
        max_columns = 2 + len( header_chain )       # +2: for unit + count columns
        add_counts = True      # ToDo
        
        logging.debug( "# of names in header_chain: %d" % len( header_chain ) )
        logging.debug( "header names in chain: %s" % str( header_chain ) )
        logging.debug( "# of names in class_chain: %d" % len( header_chain ) )
        logging.debug( "class names in chain: %s" % str( class_chain ) )
        
        nlevels = 0
        for name in header_chain:
            if name.find( "class" ) != -1:
                nlevels += 1
        logging.debug( "levels in header_chain: %d" % nlevels )
        
        # sheet_header line here; lines above for legend
        legend_offset = 9
        row = legend_offset 
        logging.debug( "# of itemchains in lex_lands: %d" % len( lex_lands ) )
        logging.debug( "lex_lands old: %s" % str( lex_lands ) )
        
        classes = list( class_chain )
        classes.sort()
        logging.debug( "class names in list: %s" % str( classes ) )
        
        logging.debug( "# of items in lex_lands_list: %d" % len( lex_lands_list ) )
        #for i, item in enumerate( lex_lands_list ):
        #    logging.debug( "%d: item: %s" % ( i, str( sorted( item, key = itemgetter( *classes ) ) ) ) )
        
        try:
            # *classes: unpack classes list as individual variables, to be used as sort keys
            lex_lands_sorted = sorted( lex_lands_list, key = itemgetter( *classes ) )
        except:
            type_, exc_value, tb = sys.exc_info()
            logging.error( "%s" % exc_value )
            lex_lands_sorted = []
        
        ndatarows = 0
        nitemchain = 0
        #for itemchain in lex_lands:
        for l, chain in enumerate( lex_lands_sorted ):
            #logging.debug( "itemchain %d: %s, %s" % ( nitemchain, itemchain, type( itemchain ) ) )
            nitemchain += 1
            #chain = json.loads( itemchain )
            logging.debug( "%d: chain: %s" %  ( l, str( chain ) ) )
            
            column = 1
            # sheet_header
            if row == legend_offset:        # table header row
                logging.debug( "# of names in header_chain: %d" % len( header_chain ) )
                logging.debug( "names in header_chain: %s" % str( header_chain ) )
                #row_name = 0
                #column_in_use = []
                
                for n, name in enumerate( sorted( header_chain ) ):
                    logging.debug( "%d: name: %s" % ( n, name ) )
                    if name == "base_year":
                        column =  1
                    elif name == "datatype":
                        column =  2
                    elif name in [ "class1", "histclass1" ]:
                        column =  3
                    elif name in [ "class2", "histclass2" ]:
                        column =  4
                    elif name in [ "class3", "histclass3" ]:
                        column =  5
                    elif name in [ "class4", "histclass4" ]:
                        column =  6
                    elif name in [ "class5", "histclass5" ]:
                        column =  7
                    elif name in [ "class6", "histclass6" ]:
                        column =  8
                    elif name in [ "class7", "histclass7" ]:
                        column =  9
                    elif name in [ "class8", "histclass8" ]:
                        column =  10
                    elif name in [ "class9", "histclass9" ]:
                        column =  11
                    elif name in [ "class10", "histclass10" ]:
                        column =  12
                    elif name == "value_unit":
                        column =  nlevels + 3
                    elif name == "count":
                        column =  nlevels + 4
                        counts_column = column
                    elif name in skip_list:
                        continue
                    else:
                        logging.debug( "name: %s ???" % name )
                        continue
                    
                    cell = ws.cell( row = row, column = column )
                    column_name = name
                    if column_name in vocab_regs_terms[ "terms" ]:
                        column_name = vocab_regs_terms[ "terms" ][ column_name ]
                    cell.value = column_name
                    logging.debug( "row: %d, header column %d: %s" % ( row, column, cell.value ) )
                
                if add_counts:
                    column += 1
                    counts_column = column
                    cell = ws.cell( row = row, column = column )
                    cell.value = "Counts"
                
                column = max_columns
                logging.debug( "# of ter_names in sorted_regions: %d" % len( sorted_regions ) )
                for ter_name in sorted_regions:
                    ter_code = regions[ ter_name ]
                    cell = ws.cell( row = row, column = column )
                    ter_name = ter_code
                    if ter_code in vocab_regs_terms[ "regions" ]:
                        ter_name = vocab_regs_terms[ "regions" ][ ter_code ]
                    cell.value = ter_name
                    column += 1
                
                row += 1
            
            # optional counts column
            add_to_counts = 0   # update total # of region counts (denominator)
            counts_row    = row
            counts_Numer  = 0                   # updated below
            counts_Denom  = len( reg_names )    # updated below
            
            # data records
            column = 1
            logging.debug( "# of names in data chain: %d" % len( chain ) )
            logging.debug( "names in chain: %s" % str( chain ) )
            try:
                base_year_chain = chain.get( "base_year" )
            except:
                base_year_chain = '0'
            
            if int( base_year ) != int( base_year_chain ):
                logging.debug( "skip: base_year %s not equal to base_year_chain %s" % ( base_year, base_year_chain ) )
                continue
            else:
                byear_counts[ str( base_year ) ] = 1 + byear_counts[ str( base_year ) ]
            
            #ter_data = lex_lands[ itemchain ]
            #lex_key = json.dumps( chain )       # this way the keys were made in preprocessor()
            
            # lex_key as stringyfied ordered dict
            keys = sorted( chain.iterkeys() )                    # sorted keys
            tuples = [ ( key, chain[ key ] ) for key in keys ]   # tuple list ordered by keys
            ordered_dict = OrderedDict( tuples )                 # dict ordered by keys
            lex_key = json.dumps( ordered_dict )                 # lex_key as string
            logging.debug( "lex_key: %s" % lex_key )
            
            ter_data = lex_lands.get( lex_key )
            if not ter_data:    # => 'na'
                logging.debug( "No ter_data from lex_key:\n%s" % lex_key )
                for l, lex_land in enumerate( lex_lands):
                    logging.debug( "%d: lex_land: %s" % ( l, lex_land ) )
                ter_data = []
            
            logging.debug( "ter_data: %s: " % str( ter_data ) )
            nclasses = 0
            db_row = chain.get( "db_row" )
            if db_row is not None:
                row = db_row + legend_offset + 1    # +1: skip table header
                counts_row = row
                logging.debug( "db_row in chain: %d" % db_row )
            
            for n, name in enumerate( sorted( chain ) ):
                logging.debug( "%d: name: %s" % ( n, name ) )
                if name == "base_year":
                    column =  1
                elif name == "datatype":
                    column =  2
                elif name in [ "class1", "histclass1" ]:
                    column =  3
                elif name in [ "class2", "histclass2" ]:
                    column =  4
                elif name in [ "class3", "histclass3" ]:
                    column =  5
                elif name in [ "class4", "histclass4" ]:
                    column =  6
                elif name in [ "class5", "histclass5" ]:
                    column =  7
                elif name in [ "class6", "histclass6" ]:
                    column =  8
                elif name in [ "class7", "histclass7" ]:
                    column =  9
                elif name in [ "class8", "histclass8" ]:
                    column =  10
                elif name in [ "class9", "histclass9" ]:
                    column =  11
                elif name in [ "class10", "histclass10" ]:
                    column =  12
                elif name == "value_unit":
                    column = nlevels + 3
                elif name == "count":
                    column =  nlevels + 4
                    counts_column = column
                    counts_value = chain[ name ]
                elif name in skip_list:
                    continue
                else:
                    logging.debug( "name: %s ???" % name )
                    continue
                
                logging.debug( "row: %d, column: %d, name: %s" % ( row, column, name ) )
                cell = ws.cell( row = row, column = column )
                value = chain[ name ]
                
                #if value == '.':
                #    value = ''
                if value == '':
                    value = '.'
                
                cell.value = value
                logging.debug( "row: %d, column: %d, name: %s, value: %s" % ( row, column, name, value ) )
            
            # display region names sorted alphabetically
            column = max_columns
            num_regions = len( sorted_regions )
            logging.debug( "number of sorted regions: %d" % num_regions )
            for idx, ter_name in enumerate( sorted_regions ):
                ter_code = regions[ ter_name ]
                if ter_code in ter_data:
                    ter_value = ter_data[ ter_code ]
                    ter_value = re.sub( r'\.0', '', str( ter_value ) )
                    try:
                        float( ter_value  )
                        counts_Numer += 1
                    except:
                        pass
                else:
                    ter_value = na
                    add_to_counts += 1
                logging.debug( "%d-of-%d ter_code: %s, ter_name: %s, ter_value: %s" % ( idx+1, num_regions, ter_code, ter_name, ter_value ) )
                
                cell = ws.cell( row = row, column = column )
                cell.value = ter_value
                #logging.debug( "%d: %s" % ( column, ter_value ) )
                column += 1
            
            logging.debug( "add_to_counts: %d, row: %d, column: %d" % ( add_to_counts, counts_row, counts_column ) )
            #if '/' in counts_value:     # update total number of regions (for given base_year)
            #if '/' in name:     # update total number of regions (for given base_year)
            cell = ws.cell( row = counts_row, column = counts_column )
            #num_denom = counts_value.split( '/' )
            #cell.value = "%s/%d" % ( num_denom[ 0 ], num_regions )
            cell.value = "%d/%d" % ( counts_Numer, counts_Denom )
            
            row += 1
            if row > legend_offset:
                ndatarows += 1
                
        nrecords.append( ndatarows )   # number of data records for current sheet
    
    logging.debug( "byear_counts: %s" % str( byear_counts ) )
    
    #logging.debug( "# of lines in sheet_header: %d" % len( sheet_header ) )
    if hist_mod == 'h':
        for l, line in enumerate( sheet_header ):
            row    = line[ "r" ] + 1
            column = line[ "c" ] + 1
            logging.debug( "row: %d, column: %d" % ( row, column ) )
            cell = ws.cell( row = row, column = column )
            cell.value = line[ "value" ]
            if l == 8:                          # update intial 0 with actual value
                cell.value = nrecords[ 0 ]      # number of data records
            logging.debug( "sheet_header l: %d, r: %d, c: %d, value: %s" % ( l, line[ "r" ], line[ "c" ], line[ "value" ] ) )
    elif hist_mod == 'm':
        for w, ws in enumerate( wb.worksheets ):
            prev_value = ""
            for l, line in enumerate( sheet_header ):
                row    = line[ "r" ] + 1
                column = line[ "c" ] + 1
                logging.debug( "row: %d, column: %d" % ( row, column ) )
                cell = ws.cell( row = row, column = column )
                if prev_value in [ "BENCHMARK-YEAR:", "ГОД:" ]:
                    cell.value = ws.title
                else:
                    cell.value = line[ "value" ]
                if l == 8:                      # update intial 0 with actual value
                    cell.value = nrecords[ w ]  # number of data records
                prev_value = cell.value
                logging.debug( "sheet_header l: %d, r: %d, c: %d, value: %s" % ( l, line[ "r" ], line[ "c" ], line[ "value" ] ) )
    
    # create copyright sheet; extract language id from filename
    comps1 = xlsx_pathname.split( '/' )
    comps2 = comps1[ -1 ].split( '-' )
    language = comps2[ 0 ]
    logging.debug( "language: %s" % language )
    
    if hist_mod == 'h':
        ws_cr = wb.create_sheet( "Copyrights", 1 )
    else:
        ws_cr = wb.create_sheet( "Copyrights", 5 )
    
    column = 1
    cell = ws_cr.cell( row = 1, column = column )
    cell.value = "_"
    cell = ws_cr.cell( row = 2, column = column )
    cell.value = "Electronic Repository of Russian Historical Statistics / Электронный архив Российской исторической статистики"
    cell = ws_cr.cell( row = 3, column = column )
    cell.value = "2014-%d" % date.today().year
    cell = ws_cr.cell( row = 4, column = column )
    cell.value = "_"
    
    if language == "en":
        cell = ws_cr.cell( row = 5, column = column )
        #cell.alignment = Alignment( horizontal = "left" ) 
        cell.value = "Creative Commons License"
        cell = ws_cr.cell( row = 6, column = column )
        cell.value = "This work is licensed under a Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License."
        cell = ws_cr.cell( row = 7, column = column )
        cell.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/"
        cell = ws_cr.cell( row = 8, column = column )
        cell.value = "_"
        cell = ws_cr.cell( row = 9, column = column )
        cell.value = "By downloading and using data from the Electronic Repository of Russian Historical Statistics the user agrees to the terms of this license. Providing a correct reference to the resource is a formal requirement of the license: "
        cell = ws_cr.cell( row = 10, column = column )
        cell.value = "Kessler, Gijs and Andrei Markevich (%d), Electronic Repository of Russian Historical Statistics, 18th - 21st centuries, http://ristat.org/" % date.today().year
    elif language == "ru":
        cell = ws_cr.cell( row = 5, column = column )
        cell.value = "Лицензия Creative Commons"
        cell = ws_cr.cell( row = 6, column = column )
        cell.value = "Это произведение доступно по лицензии Creative Commons «Attribution-NonCommercial-ShareAlike» («Атрибуция — Некоммерческое использование — На тех же условиях») 4.0 Всемирная."
        cell = ws_cr.cell( row = 7, column = column )
        cell.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/deed.ru"
        cell = ws_cr.cell( row = 8, column = column )
        cell.value = "_"
        cell = ws_cr.cell( row = 9, column = column )
        cell.value = "Скачивая и начиная использовать данные пользователь автоматически соглашается с этой лицензией. Наличие корректно оформленной ссылки является обязательным требованием лицензии:"
        cell = ws_cr.cell( row = 10, column = column )
        cell.value = "Кесслер Хайс и Маркевич Андрей (%d), Электронный архив Российской исторической статистики, XVIII – XXI вв., [Электронный ресурс] : [сайт]. — Режим доступа: http://ristat.org/" % date.today().year

    try:
        wb.save( xlsx_pathname )
        msg = None
    except:
        type_, value, tb = exc_info()
        msg = "saving xlsx failed: %s" % value
    
    # sigh, openpyxl can't sort, let's sort with pandas
    
    params = {
        "lang"       : lang,
        "hist_mod"   : hist_mod, 
        "topic_name" : topic_name, 
        "base_year"  : base_year
    }
    
    # sorting now via temp postgres table
    #pandas_sort( xlsx_pathname, nlevels, params )
    
    return xlsx_pathname, msg



def aggregate_dataset_records( key, download_dir, xlsx_name, lex_lands, vocab_regs_terms, sheet_header, topic_name, params ):
    logging.info( "aggregate_dataset_records()" )
    logging.debug( "key: %s" % key )
    logging.debug( "download_dir: %s" % download_dir )
    logging.debug( "xlsx_name: %s" % xlsx_name )
    logging.debug( "lex_lands: %s" % str( lex_lands ) )
    logging.debug( "vocab_regs_terms: %s" % str( vocab_regs_terms ) )
    logging.debug( "sheet_header type: %s" % type( sheet_header ) )
    logging.debug( "topic_name: %s" % topic_name )
    logging.debug( "params: %s" % str( params ) )
    
    xlsx_pathname = ""
    if not os.path.isdir( download_dir ):
        msg = "download destination was removed"
        return xlsx_pathname, msg

    xlsx_pathname = os.path.abspath( os.path.join( download_dir, xlsx_name ) )
    logging.debug( "full_path: %s" % xlsx_pathname )
    
    lang = ""
    hist_mod = ""
    base_year_str = ""
    key_comps = key.split( '-' )
    if len( key_comps ) >= 1:
        lang = key_comps[ 0 ]
    if len( key_comps ) >= 2:
        hist_mod = key_comps[ 1 ]
    if len( key_comps ) >= 3:
        base_year_str = key_comps[ 3 ]
        base_year = int( base_year_str )
    logging.debug( "lang: %s, hist_mod: %s, base_year_str: %s" % ( lang, hist_mod, base_year_str ) )
    
    logging.debug( "# of keys in vocab_regs_terms: %d" % len( vocab_regs_terms.keys() ) )
    for k, v in vocab_regs_terms.iteritems():
        logging.debug( "k: %s, v: %s" % ( k, str( v ) ) )
    
    nsheets = 0
    ter_code_list = []
    
    #wb = openpyxl.Workbook( encoding = "utf-8" )
    wb = openpyxl.Workbook()
    
    escape = False
    if escape:
        logging.debug( "openpyxl version: %s" % openpyxl.__version__ )
        logging.debug( "empty spreadsheet" )
        ws = wb.active                  # grab the active worksheet
        ws['A1'] = 42                   # Data can be assigned directly to cells
        ws.append([1, 2, 3])            # Rows can also be appended
        import datetime                 # Python types will automatically be converted
        ws['A2'] = datetime.datetime.now()
        wb.save( "/home/dpe/tmp/data/download/sample.xlsx" )
        wb.save( xlsx_pathname )    # Save the file
        return xlsx_pathname, ""
    
    clientcache = MongoClient()
    db_cache = clientcache.get_database( 'datacache' )
    
    #logging.debug( "find cached data with key: %s" % key )
    #result = db_cache.data.find( { "key": key } )
    #logging.info( "aggregate_dataset_records() length of cached dict: %d for key: %s" % ( len( result ), key ) )
    
    db_vocabulary = clientcache.get_database( 'vocabulary' )   # vocabulary
    
    logging.debug( "params: %s, %s" % ( params, type( params ) ) )
    if params:
        for k in params:
            logging.debug( "key: %s, value: %s" % ( k, params[ k ] ) )
    
        try:
            ter_code_list = params[ "ter_codes" ]
        except:
            ter_code_list = []
    
        logging.debug( "ter_code_list: %s, %s" % ( str( ter_code_list ), type( ter_code_list ) ) )
        
        try:
            level_paths = params[ "path" ]
        except:
            level_paths = []
        
        logging.debug( "# of level paths: %d" % len( level_paths ) )
        for level_path in level_paths:
            logging.debug( "level path: %s" % level_path )
    else:
        logging.error( "aggregate_dataset_records() NO params" )
    
    nrecords = []       # number of data records per sheet
    if hist_mod == 'h':
        nsheets = 1
        base_years = [ base_year ]
        ws_0 = wb.get_active_sheet()
        ws_0.title = "Dataset"

    elif hist_mod == 'm':
        nsheets = 5
        base_years = [ "1795", "1858", "1897", "1959", "2002" ]
        ws_0 = wb.get_active_sheet()
        ws_0.title = "1795"
        ws_1 = wb.create_sheet( "1858", 1 )
        ws_2 = wb.create_sheet( "1897", 2 )
        ws_3 = wb.create_sheet( "1959", 3 )
        ws_4 = wb.create_sheet( "2002", 4 )
        
        vocab_regions = db_vocabulary.data.find( { "vocabulary": "ERRHS_Vocabulary_regions" } )
        #logging.debug( "vocab_regions: %s" % vocab_regions )
        ter_code_list = []
        for region in vocab_regions:
            logging.debug( "region: %s" % str( region ) )
            ter_code_list.append( region[ "ID" ] )
    
    # initialise base_year counts
    byear_counts = { "1795" : 0, "1858" : 0, "1897" : 0, "1959" : 0, "2002" : 0 }
    
    #skip_list = [ "count" ]
    skip_list = []
    
    # loop over the data sheets, and select the correct year data
    for sheet_idx in range( nsheets ):
        base_year = base_years[ sheet_idx ]
        logging.debug( "work_sheet: %d, base_year: %s" % ( sheet_idx, base_year ) )
        if sheet_idx == 0:
            ws = ws_0
        elif sheet_idx == 1:
            ws = ws_1
        elif sheet_idx == 2:
            ws = ws_2
        elif sheet_idx == 3:
            ws = ws_3
        elif sheet_idx == 4:
            ws = ws_4
        
        # load regions for base_year
        regions_filter = {}
        regions_filter[ "vocabulary" ] = "ERRHS_Vocabulary_regions"
        # regions depend on base_year
        regions_filter[ 'basisyear' ] = str( base_year )    # basisyear is int column in vocabulary file
        logging.debug( "regions_filter: %s" % str( regions_filter ) )
        cursor = db_vocabulary.data.find( regions_filter )
        logging.debug( "# of regions from vocabulary: %d" % cursor.count() )
        vocab_regions = list( cursor )
        
        # get region names from ter_codes
        regions = {}
        #ter_codes = vocab_regs_terms[ "ter_codes" ]    # from sql result
        ter_codes = ter_code_list                       # from qinput or all (modern)
        logging.debug( "ter_codes: %s" % str( ter_codes ) )
        for i, item in enumerate( vocab_regions ):
            logging.debug( "%d: %s" % ( i, str( item ) ) )
            ID = item[ 'ID' ]
            if ID in ter_codes:
                if lang == 'en':
                    regions[ item[ 'ID' ] ] = item[ 'EN' ]
                else:
                    regions[ item[ 'ID' ] ] = item[ 'RUS' ]
        
        logging.debug( "regions: %s" % str( regions ) )
        vocab_regs_terms[ "regions" ] = regions
        
        logging.debug( "%d keys in vocab_regs_terms:" % len( vocab_regs_terms ) )
        for key in vocab_regs_terms:
            logging.debug( "key: %s" % key )
        
        try:
            na = vocab_regs_terms[ "terms" ][ "na" ]
        except:
            na = "na"
        logging.debug( "na: %s" % na )
        
        regions = {}        # key: reg_name, value: reg_code
        reg_names = []      # list of reg_name's
        for ter_code in vocab_regs_terms[ "regions" ]:
            tmpname = vocab_regs_terms[ "regions" ][ ter_code ]
            ter_name = tmpname.decode( 'utf-8' )
            regions[ ter_name ] = ter_code
            reg_names.append( ter_name ) 
            
        logging.debug( "unsorted reg_names: %s" % str( reg_names ) )
        # sort the region names alphabetically
        locale = Locale( 'ru' )  # 'el' is the locale code for Greek
        collator = Collator.createInstance( locale )
        sorted_regions = sorted( reg_names, cmp = collator.compare )
        logging.debug( "sorted_regions: %s" % str( sorted_regions ) )
        
        # empty level strings are not in the data, but we need same # of colums 
        # for all rows; find the maximum number of levels
        header_chain = set()
        class_chain  = set()
        lex_lands_list0 = []
        for itemchain in lex_lands:
            chain = json.loads( itemchain )
            lex_lands_list0.append( chain )
            
            for name in sorted( chain ):
                header_chain.add( name )
                if "class" in name:
                    class_chain.add( name )
        
        lex_lands_list = []
        for chain in lex_lands_list0:
            for class_ in class_chain:
                if not chain.get( class_ ):
                    chain[ class_ ] = ''
            lex_lands_list.append( chain )
            logging.debug( "chain: %s" % str( chain ) )
        
        #max_columns = len( header_chain )
        #max_columns = 1 + len( header_chain )       # +1: for count column
        max_columns = 2 + len( header_chain )       # +2: for unit + count columns
        add_counts = True      # ToDo
        
        logging.debug( "# of names in header_chain: %d" % len( header_chain ) )
        logging.debug( "header names in chain: %s" % str( header_chain ) )
        logging.debug( "# of names in class_chain: %d" % len( header_chain ) )
        logging.debug( "class names in chain: %s" % str( class_chain ) )
        
        nlevels = 0
        for name in header_chain:
            if name.find( "class" ) != -1:
                nlevels += 1
        logging.debug( "levels in header_chain: %d" % nlevels )
        
        # sheet_header line here; lines above for legend
        legend_offset = 9
        row = legend_offset 
        logging.debug( "# of itemchains in lex_lands: %d" % len( lex_lands ) )
        logging.debug( "lex_lands old: %s" % str( lex_lands ) )
        
        classes = list( class_chain )
        classes.sort()
        logging.debug( "class names in list: %s" % str( classes ) )
        
        logging.debug( "# of items in lex_lands_list: %d" % len( lex_lands_list ) )
        #for i, item in enumerate( lex_lands_list ):
        #    logging.debug( "%d: item: %s" % ( i, str( sorted( item, key = itemgetter( *classes ) ) ) ) )
        
        try:
            # *classes: unpack classes list as individual variables, to be used as sort keys
            lex_lands_sorted = sorted( lex_lands_list, key = itemgetter( *classes ) )
        except:
            type_, exc_value, tb = sys.exc_info()
            logging.error( "%s" % exc_value )
            lex_lands_sorted = []
        
        ndatarows = 0
        nitemchain = 0
        #for itemchain in lex_lands:
        for l, chain in enumerate( lex_lands_sorted ):
            #logging.debug( "itemchain %d: %s, %s" % ( nitemchain, itemchain, type( itemchain ) ) )
            nitemchain += 1
            #chain = json.loads( itemchain )
            logging.debug( "%d: chain: %s" %  ( l, str( chain ) ) )
            
            column = 1
            # sheet_header
            if row == legend_offset:        # table header row
                logging.debug( "# of names in header_chain: %d" % len( header_chain ) )
                logging.debug( "names in header_chain: %s" % str( header_chain ) )
                #row_name = 0
                #column_in_use = []
                
                for n, name in enumerate( sorted( header_chain ) ):
                    logging.debug( "%d: name: %s" % ( n, name ) )
                    if name == "base_year":
                        column =  1
                    elif name == "datatype":
                        column =  2
                    elif name in [ "class1", "histclass1" ]:
                        column =  3
                    elif name in [ "class2", "histclass2" ]:
                        column =  4
                    elif name in [ "class3", "histclass3" ]:
                        column =  5
                    elif name in [ "class4", "histclass4" ]:
                        column =  6
                    elif name in [ "class5", "histclass5" ]:
                        column =  7
                    elif name in [ "class6", "histclass6" ]:
                        column =  8
                    elif name in [ "class7", "histclass7" ]:
                        column =  9
                    elif name in [ "class8", "histclass8" ]:
                        column =  10
                    elif name in [ "class9", "histclass9" ]:
                        column =  11
                    elif name in [ "class10", "histclass10" ]:
                        column =  12
                    elif name == "value_unit":
                        column =  nlevels + 3
                    elif name == "count":
                        column =  nlevels + 4
                        counts_column = column
                    elif name == "ter_codes":
                        #pass
                        continue
                    elif name in skip_list:
                        continue
                    else:
                        logging.debug( "name: %s ???" % name )
                        continue
                    
                    """
                    if name == "ter_codes":
                        for ter_code_str in ter_codes:
                            logging.debug( "ter_code_str: %s" % ( ter_code_str ) )
                            ter_code_dict = json.loads( ter_code_str )
                            ter_code = ter_code_dict.get( "ter_code" )
                            total    = ter_code_dict.get( "total" )
                            logging.debug( "ter_code: %s, total: %s" % ( ter_code, total ) )
                    """
                    
                    cell = ws.cell( row = row, column = column )
                    column_name = name
                    if column_name in vocab_regs_terms[ "terms" ]:
                        column_name = vocab_regs_terms[ "terms" ][ column_name ]
                    cell.value = column_name
                    logging.debug( "row: %d, header column %d: %s" % ( row, column, cell.value ) )
                
                if add_counts:
                    column += 1
                    counts_column = column
                    cell = ws.cell( row = row, column = column )
                    cell.value = "Counts"
                
                column = max_columns
                logging.debug( "# of ter_names in sorted_regions: %d" % len( sorted_regions ) )
                for ter_name in sorted_regions:
                    ter_code = regions[ ter_name ]
                    cell = ws.cell( row = row, column = column )
                    ter_name = ter_code
                    if ter_code in vocab_regs_terms[ "regions" ]:
                        ter_name = vocab_regs_terms[ "regions" ][ ter_code ]
                    cell.value = ter_name
                    column += 1
                
                row += 1
            
            # optional counts column
            add_to_counts = 0   # update total # of region counts (denominator)
            counts_row    = row
            counts_Numer  = 0                   # updated below
            counts_Denom  = len( reg_names )    # updated below
            
            # data records
            column = 1
            logging.debug( "# of names in data chain: %d" % len( chain ) )
            logging.debug( "names in chain: %s" % str( chain ) )
            
            try:
                base_year_chain = chain.get( "base_year" )
            except:
                base_year_chain = '0'
            logging.info( "base_year_chain: %s" % base_year_chain )
            
            # 29-Jan-2019 No base_year
            logging.info( "params: %s" % str( params ) )
            logging.info( "hist_mod: %s" % hist_mod )
            if hist_mod == 'h' and base_year is None:
                base_year = params.get( "base_year" )
            logging.info( "base_year: %s" % base_year )
            
            # hack
            if base_year_chain is None:
                base_year_chain = base_year
            
            if int( base_year ) != int( base_year_chain ):
                logging.debug( "skip: base_year %s not equal to base_year_chain %s" % ( base_year, base_year_chain ) )
                continue
            else:
                byear_counts[ str( base_year ) ] = 1 + byear_counts[ str( base_year ) ]
            
            #ter_data = lex_lands[ itemchain ]
            #lex_key = json.dumps( chain )       # this way the keys were made in preprocessor()
            
            # lex_key as stringyfied ordered dict
            keys = sorted( chain.iterkeys() )                    # sorted keys
            tuples = [ ( key, chain[ key ] ) for key in keys ]   # tuple list ordered by keys
            ordered_dict = OrderedDict( tuples )                 # dict ordered by keys
            lex_key = json.dumps( ordered_dict )                 # lex_key as string
            logging.debug( "lex_key: %s" % lex_key )
            
            ter_data = lex_lands.get( lex_key )
            if not ter_data:    # => 'na'
                logging.debug( "No ter_data from lex_key:\n%s" % lex_key )
                for l, lex_land in enumerate( lex_lands):
                    logging.debug( "%d: lex_land: %s" % ( l, lex_land ) )
                ter_data = []
            
            logging.debug( "ter_data: %s: " % str( ter_data ) )
            nclasses = 0
            db_row = chain.get( "db_row" )
            if db_row is not None:
                row = db_row + legend_offset + 1    # +1: skip table header
                counts_row = row
                logging.debug( "db_row in chain: %d" % db_row )
            
            for n, name in enumerate( sorted( chain ) ):
                logging.debug( "%d: name: %s" % ( n, name ) )
                if name == "base_year":
                    column =  1
                elif name == "datatype":
                    column =  2
                elif name in [ "class1", "histclass1" ]:
                    column =  3
                elif name in [ "class2", "histclass2" ]:
                    column =  4
                elif name in [ "class3", "histclass3" ]:
                    column =  5
                elif name in [ "class4", "histclass4" ]:
                    column =  6
                elif name in [ "class5", "histclass5" ]:
                    column =  7
                elif name in [ "class6", "histclass6" ]:
                    column =  8
                elif name in [ "class7", "histclass7" ]:
                    column =  9
                elif name in [ "class8", "histclass8" ]:
                    column =  10
                elif name in [ "class9", "histclass9" ]:
                    column =  11
                elif name in [ "class10", "histclass10" ]:
                    column =  12
                elif name == "value_unit":
                    column = nlevels + 3
                elif name == "count":
                    column =  nlevels + 4
                    counts_column = column
                    counts_value = chain[ name ]
                elif name in skip_list:
                    continue
                else:
                    logging.debug( "name: %s ???" % name )
                    continue
                
                logging.debug( "row: %d, column: %d, name: %s" % ( row, column, name ) )
                cell = ws.cell( row = row, column = column )
                value = chain[ name ]
                
                #if value == '.':
                #    value = ''
                if value == '':
                    value = '.'
                
                cell.value = value
                logging.debug( "row: %d, column: %d, name: %s, value: %s" % ( row, column, name, value ) )
            
            # display region names sorted alphabetically
            column = max_columns
            num_regions = len( sorted_regions )
            logging.debug( "number of sorted regions: %d" % num_regions )
            for idx, ter_name in enumerate( sorted_regions ):
                ter_code = regions[ ter_name ]
                if ter_code in ter_data:
                    ter_value = ter_data[ ter_code ]
                    ter_value = re.sub( r'\.0', '', str( ter_value ) )
                    try:
                        float( ter_value  )
                        counts_Numer += 1
                    except:
                        pass
                else:
                    ter_value = na
                    add_to_counts += 1
                logging.debug( "%d-of-%d ter_code: %s, ter_name: %s, ter_value: %s" % ( idx+1, num_regions, ter_code, ter_name, ter_value ) )
                
                cell = ws.cell( row = row, column = column )
                cell.value = ter_value
                #logging.debug( "%d: %s" % ( column, ter_value ) )
                column += 1
            
            logging.debug( "add_to_counts: %d, row: %d, column: %d" % ( add_to_counts, counts_row, counts_column ) )
            #if '/' in counts_value:     # update total number of regions (for given base_year)
            #if '/' in name:     # update total number of regions (for given base_year)
            cell = ws.cell( row = counts_row, column = counts_column )
            #num_denom = counts_value.split( '/' )
            #cell.value = "%s/%d" % ( num_denom[ 0 ], num_regions )
            cell.value = "%d/%d" % ( counts_Numer, counts_Denom )
            
            row += 1
            if row > legend_offset:
                ndatarows += 1
                
        nrecords.append( ndatarows )   # number of data records for current sheet
    
    logging.debug( "byear_counts: %s" % str( byear_counts ) )
    
    #logging.debug( "# of lines in sheet_header: %d" % len( sheet_header ) )
    if hist_mod == 'h':
        for l, line in enumerate( sheet_header ):
            row    = line[ "r" ] + 1
            column = line[ "c" ] + 1
            logging.debug( "row: %d, column: %d" % ( row, column ) )
            cell = ws.cell( row = row, column = column )
            cell.value = line[ "value" ]
            if l == 8:                          # update intial 0 with actual value
                cell.value = nrecords[ 0 ]      # number of data records
            logging.debug( "sheet_header l: %d, r: %d, c: %d, value: %s" % ( l, line[ "r" ], line[ "c" ], line[ "value" ] ) )
    elif hist_mod == 'm':
        for w, ws in enumerate( wb.worksheets ):
            prev_value = ""
            for l, line in enumerate( sheet_header ):
                row    = line[ "r" ] + 1
                column = line[ "c" ] + 1
                logging.debug( "row: %d, column: %d" % ( row, column ) )
                cell = ws.cell( row = row, column = column )
                if prev_value in [ "BENCHMARK-YEAR:", "ГОД:" ]:
                    cell.value = ws.title
                else:
                    cell.value = line[ "value" ]
                if l == 8:                      # update intial 0 with actual value
                    cell.value = nrecords[ w ]  # number of data records
                prev_value = cell.value
                logging.debug( "sheet_header l: %d, r: %d, c: %d, value: %s" % ( l, line[ "r" ], line[ "c" ], line[ "value" ] ) )
    
    # create copyright sheet; extract language id from filename
    comps1 = xlsx_pathname.split( '/' )
    comps2 = comps1[ -1 ].split( '-' )
    language = comps2[ 0 ]
    logging.debug( "language: %s" % language )
    
    if hist_mod == 'h':
        ws_cr = wb.create_sheet( "Copyrights", 1 )
    else:
        ws_cr = wb.create_sheet( "Copyrights", 5 )
    
    column = 1
    cell = ws_cr.cell( row = 1, column = column )
    cell.value = "_"
    cell = ws_cr.cell( row = 2, column = column )
    cell.value = "Electronic Repository of Russian Historical Statistics / Электронный архив Российской исторической статистики"
    cell = ws_cr.cell( row = 3, column = column )
    cell.value = "2014-%d" % date.today().year
    cell = ws_cr.cell( row = 4, column = column )
    cell.value = "_"
    
    if language == "en":
        cell = ws_cr.cell( row = 5, column = column )
        #cell.alignment = Alignment( horizontal = "left" ) 
        cell.value = "Creative Commons License"
        cell = ws_cr.cell( row = 6, column = column )
        cell.value = "This work is licensed under a Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License."
        cell = ws_cr.cell( row = 7, column = column )
        cell.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/"
        cell = ws_cr.cell( row = 8, column = column )
        cell.value = "_"
        cell = ws_cr.cell( row = 9, column = column )
        cell.value = "By downloading and using data from the Electronic Repository of Russian Historical Statistics the user agrees to the terms of this license. Providing a correct reference to the resource is a formal requirement of the license: "
        cell = ws_cr.cell( row = 10, column = column )
        cell.value = "Kessler, Gijs and Andrei Markevich (%d), Electronic Repository of Russian Historical Statistics, 18th - 21st centuries, http://ristat.org/" % date.today().year
    elif language == "ru":
        cell = ws_cr.cell( row = 5, column = column )
        cell.value = "Лицензия Creative Commons"
        cell = ws_cr.cell( row = 6, column = column )
        cell.value = "Это произведение доступно по лицензии Creative Commons «Attribution-NonCommercial-ShareAlike» («Атрибуция — Некоммерческое использование — На тех же условиях») 4.0 Всемирная."
        cell = ws_cr.cell( row = 7, column = column )
        cell.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/deed.ru"
        cell = ws_cr.cell( row = 8, column = column )
        cell.value = "_"
        cell = ws_cr.cell( row = 9, column = column )
        cell.value = "Скачивая и начиная использовать данные пользователь автоматически соглашается с этой лицензией. Наличие корректно оформленной ссылки является обязательным требованием лицензии:"
        cell = ws_cr.cell( row = 10, column = column )
        cell.value = "Кесслер Хайс и Маркевич Андрей (%d), Электронный архив Российской исторической статистики, XVIII – XXI вв., [Электронный ресурс] : [сайт]. — Режим доступа: http://ristat.org/" % date.today().year

    try:
        wb.save( xlsx_pathname )
        msg = None
    except:
        type_, value, tb = exc_info()
        msg = "saving xlsx failed: %s" % value
    
    # sigh, openpyxl can't sort, let's sort with pandas
    
    params = {
        "lang"       : lang,
        "hist_mod"   : hist_mod, 
        "topic_name" : topic_name, 
        "base_year"  : base_year
    }
    
    # sorting now via temp postgres table
    #pandas_sort( xlsx_pathname, nlevels, params )
    
    return xlsx_pathname, msg



def pandas_sort( xlsx_pathname_in, nlevels, params ):
    logging.info( "pandas_sort()" )
    
    ( xlsx_head, xlsx_tail ) = os.path.split( xlsx_pathname_in )
    ( xlsx_root, xlsx_ext ) = os.path.splitext( xlsx_tail )
    xlsx_tail_pd = xlsx_root + "-pd" + xlsx_ext
    xlsx_pathname_out = os.path.join( xlsx_head, xlsx_tail_pd )

    logging.debug( "xlsx_pathname_in: %s" % xlsx_pathname_in )
    logging.debug( "xlsx_head: %s" % xlsx_head )
    logging.debug( "xlsx_tail: %s" % xlsx_tail )
    logging.debug( "xlsx_root: %s" % xlsx_root )
    logging.debug( "xlsx_ext: %s" % xlsx_ext )
    logging.debug( "xlsx_pathname_out: %s" % xlsx_pathname_out )

    # sort dataframe rows by Levels ([hist]classes)
    sort_levels = []
    for l in range( nlevels ):
        l_str = "Level %d" % ( l + 1 )
        sort_levels.append( l_str )
    
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    #engine = "openpyxl"
    engine = "xlsxwriter"
    writer = pd.ExcelWriter( xlsx_pathname_out, engine = engine )
    
    #EF = pd.read_excel( xlsx_pathname_in )  # AttributeError: 'DataFrame' object has no attribute 'sheet_names'
    EF = pd.ExcelFile( xlsx_pathname_in )
    
    for sheet_name in EF.sheet_names:
        logging.info( "sheet: %s" % sheet_name )
        
        # index = False: no row numbers in output (but they did show the sorting effect)
        if sheet_name == "Copyrights":
            nskiprows = 0
            df = EF.parse( sheet_name )
            df.to_excel( writer, sheet_name = sheet_name, index = False )
        else:
            nskiprows = 8       # start at table header
            df_table = EF.parse( sheet_name, skiprows = nskiprows )
            
            df = pd.DataFrame()
            df = df_table
            
            # sort the table
            try:
                df = df.sort_values( by = sort_levels )
            except:
                logging.error( "sorting failed with  with sort_levels:\n%s" % sort_levels )
                type_, value, tb = sys.exc_info()
                logging.error( "%s" % value )
            
            # set comment lines below the table
            header = list( df )
            c0 = header[ 0 ]
            c1 = header[ 1 ]
            nrows = len( df )
            row_offset = nrows
            
            lang       = params[ "lang" ]
            hist_mod   = params[ "hist_mod" ]
            topic_name = params[ "topic_name" ]
            base_year  = params[ "base_year" ]
            
            df.loc[ row_offset, c0 ] = ""
            df.loc[ row_offset, c1 ] = ""
                
            if lang == 'en':
                if hist_mod == 'h':
                    classification = "HISTORICAL"
                elif hist_mod == 'm':
                    classification = "MODERN"
                else:
                    classification = ""
                
                df.loc[ row_offset + 1, c0 ] = "Electronic repository of Russian Historical Statistics - ristat.org"
                df.loc[ row_offset + 3, c0 ] = "TOPIC:"
                df.loc[ row_offset + 3, c1 ] = topic_name
                df.loc[ row_offset + 4, c0 ] = "BENCHMARK-YEAR:"
                df.loc[ row_offset + 4, c1 ] = base_year
                df.loc[ row_offset + 5, c0 ] = "CLASSIFICATION:"
                df.loc[ row_offset + 5, c1 ] = classification
                df.loc[ row_offset + 6, c0 ] = "NUMBER"
                df.loc[ row_offset + 6, c1 ] = nrows
            else:
                if hist_mod == 'h':
                    classification = "ИСТОРИЧЕСКАЯ"
                elif hist_mod == 'm':
                    classification = "СОВРЕМЕННАЯ"
                else:
                    classification = ""
                
                df.loc[ row_offset + 1, c0 ] = "Электронный архив Российской исторической статистики - ristat.org"
                df.loc[ row_offset + 3, c0 ] = "ТЕМА:"
                df.loc[ row_offset + 3, c1 ] = topic_name
                df.loc[ row_offset + 4, c0 ] = "ГОД:"
                df.loc[ row_offset + 4, c1 ] = base_year
                df.loc[ row_offset + 5, c0 ] = "КЛАССИФИКАЦИЯ:"
                df.loc[ row_offset + 5, c1 ] = classification
                df.loc[ row_offset + 6, c0 ] = "ЧИСЛО"
                df.loc[ row_offset + 6, c1 ] = nrows
            
            df.to_excel( writer, sheet_name = sheet_name, index = False )
        
    writer.save()

# [eof]
