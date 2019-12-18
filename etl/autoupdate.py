#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
This script retrieves RiStat files from Dataverse and updates MongoDB. 
- vocabulary *.tab files are stored locally and/or in MongoDB
- gets vocabulary data from PostgreSQL and stores it MongoDB
- transforms binary xlsx spreadsheet files to csv files


VT-07-Jul-2016 latest change by VT
FL-03-Mar-2017 Py2/Py3 compatibility: using pandas instead of xlsx2csv to create csv files
FL-03-Mar-2017 Py2/Py3 compatibility: using future-0.16.0
FL-27-Mar-2017 Also download documentation files
FL-17-May-2017 postgresql datasets.topics counts
FL-03-Jul-2017 Translate data files to english
FL-07-Jul-2017 sys.stderr.write() cannot write to cron.log as normal user
FL-11-Jul-2017 pandas: do not parse numbers, but keep strings as they are
FL-13-Aug-2017 Py2/Py3 cleanup
FL-18-Dec-2017 Keep trailing input '\n' for header lines in translate_csv
FL-16-Jan-2018 Separate RU & EN tables
FL-15-May-2018 Rounding of data in value column
FL-23-Jul-2018 DATATYPE 1.10 => 1.1 ? try to suppress unwanted conversion of DATATYPE
FL-18-Dec-2018 Rounding of float values in filter_csv() was not active
FL-06-Mar-2019 HISTCLASS1 & CLASS2 digits strip spurious .0 float ending
FL-21-Mar-2019 Prepare downloads for filecatalogue
FL-21-Mar-2019 PostgreSQL increase max_connections
FL-28-Mar-2019 downloads for filecatalogue
FL-08-Jun-2019 openpyxl for xslx to csv coversion
FL-08-Jun-2019 from backports import csv; openpyxl helper
FL-24-Jun-2019 document info request from GUI
FL-02-Jul-2019 use requests instead of urllib[2]
FL-04-Jul-2019 autoupdate steered by dataverse
FL-29-Jul-2019 new ristat-key, but failed!
FL-06-Aug-2019 AUTOUPDATE > 1 : force doing an autoupdate
FL-19-Nov-2019 Separate retrieving documents from processing documents
FL-03-Dec-2019 Now using tablib for excel processing
FL-09-Dec-2019 Update territory names in xlsx2csv_tablib_filter()
FL-09-Dec-2019 Eliminate now redundant filtereing before postgres insert
FL-17-Dec-2019 AUTOUPDATE bug
FL-18-Dec-2019 

TODO
- Use pyDataverse from PyPI

--------------------------------------------------------------------------------
#def load_json( apiurl ):
def empty_dir( dst_dir ):
def documents_by_handle( config_parser, handle_name, dst_dir, dv_format = "", check_ts = False ):
def loadjson( json_dataurl ):
def documents_info( config_parser, language ):
def update_documentation( config_parser ):
def check_autoupdate( config_parser, dv_format ):
def retrieve_vocabularies( config_parser, dv_format, check_ts = False ):
def copy_doc_src2dst():
def convert_vocabularies2csv( convert_vocabularies ):
def merge_vocabs( vocab_csv_dir ):
def mongo_store_vocabularies():
def retrieve_handle_docs( config_parser, handle_name, dv_format = "" ):
def row_count( config_parser, language ):
def clear_postgres_tale( config_parser, language ):
def store_handle_docs( config_parser, handle_name, language ):
def test_csv_file( path_name ):
def filter_csv( csv_dir, in_filename ):
def update_handle_docs( config_parser, mongo_client ):
def clear_mongo( mongo_client ):
def topic_counts( config_parser, language ):
def load_vocab( config_parser, vocab_fname, vocab, pos_rus, pos_eng ):
def convert_excel2csv( config_parser, excel_package ):
def xlsx2csv_pandas( xlsx_dir, xlsx_filename, csv_dir, extra ):
def xlsx2csv_openpyxl( xlsx_dir, xlsx_filename, csv_dir, extra ):
def xlsx2csv_tablib( xlsx_dir, xlsx_filename, csv_dir, extra ):
def xlsx2csv_tablib_filter( vocab_regions, vocab_units_ru, xlsx_dir, xlsx_filename, csv_dir, extra ):
def translate_errhs_csvs( config_parser, handle_names ):
def translate_csv( config_parser, handle_name, vocab_units, vocab_regions, vocab_histclasses, vocab_modclasses ):
def compile_filecatalogue( config_parser, language, excel_package, pd_engine ):
def csv2xlsx_pandas( language, vocab_units, csv_dir, csv_filename, xlsx_dir ):
def csv2xlsx_tablib( language, vocab_units, csv_dir, csv_filename, xlsx_dir ):
def format_secs( seconds ):
"""


# future-0.18.2 imports for Python 2/3 compatibility
from __future__ import ( absolute_import, division, print_function, unicode_literals )
from builtins import ( ascii, bytes, chr, dict, filter, hex, input, int, list, map, 
    next, object, oct, open, pow, range, round, super, str, zip )

from six.moves import configparser, StringIO

import codecs
import ConfigParser
import csv
import dateutil
import getpass
import io
import json
import logging
import math
import openpyxl
import os
import pandas as pd
import psycopg2
import re
import requests
import simplejson
import sys
import shutil
import tablib

from backports import csv
from bidict import bidict
from datetime import date, datetime
from pymongo import MongoClient
from shutil import copy2
from sys import exc_info
from time import ctime, mktime, time

sys.path.insert( 0, os.path.abspath( os.path.join(os.path.dirname( "__file__" ), './' ) ) )
sys.path.insert( 0, os.path.abspath( os.path.join(os.path.dirname( "__file__" ), '../' ) ) )
sys.path.insert( 0, os.path.abspath( os.path.join(os.path.dirname( "__file__" ), '../service' ) ) )
#print( sys.path )
#print( "pwd:", os.getcwd() )

from dataverse import Connection

from vocab import vocabulary, classupdate
from service.configutils import DataFilter

autoupdate = False
Nexcept = 0

# column comment_source of postgresql table russianrepository of db ristat
COMMENT_LENGTH_MAX_DB = 4096
# primary key for russianrepository table
pkey = None

#"""
handle_names = [ 
    "hdl_errhs_population",     # ERRHS_1   40 files
    "hdl_errhs_capital",        # 
    "hdl_errhs_industry",       # 
    "hdl_errhs_agriculture",    # ERRHS_4   10 files
    "hdl_errhs_labour",         # ERRHS_2   12 files
    "hdl_errhs_services",       # 
    "hdl_errhs_land"            # ERRHS_7    2 files
]
#"""
#handle_names = [ "hdl_errhs_agriculture" ]     # test for rounding
#handle_names = [ "hdl_errhs_land" ]     		# test for rounding


def get_configparser():
    RUSSIANREPO_CONFIG_PATH = os.environ[ "RUSSIANREPO_CONFIG_PATH" ]
    logging.debug( "RUSSIANREPO_CONFIG_PATH: %s" % RUSSIANREPO_CONFIG_PATH )
    
    configpath = RUSSIANREPO_CONFIG_PATH
    if not os.path.isfile( configpath ):
        print( "in %s" % __file__ )
        print( "configpath %s FILE DOES NOT EXIST" % configpath )
        print( "EXIT" )
        sys.exit( 1 )
    
    logging.debug( "configpath: %s" % configpath )
    
    configparser = ConfigParser.RawConfigParser()
    configparser.read( configpath )
    
    return configparser
# get_configparser()


def empty_dir( dst_dir ):
    global Nexcept
    
    logging.info( "empty_dir() %s" % dst_dir )
    logging.info( "removing previous downloads" )
    
    for root, sdirs, files in os.walk( dst_dir ):
        #logging.info( "root:  %s" % str( root ) )
        #logging.info( "sdirs: %s" % str( sdirs ) )
        #logging.info( "files: %s" % str( files ) )
        
        if files is not None:
            files.sort()
        for fname in files:
            file_path = os.path.join( root, fname )
            # only delete files we recognize
            if fname.startswith( "ERRHS_" ):
                mtime = os.path.getmtime( file_path )
                timestamp = ctime( mtime )
                logging.debug( "removing file: (created: %s) %s" % ( timestamp, file_path ) )
                try:
                    os.unlink( file_path )
                except:
                    Nexcept += 1
                    type_, value, tb = sys.exc_info()
                    logging.error( "%s" % value )
        
        if sdirs is not None:
            sdirs.sort()
        for dname in sdirs:
            # only delete dirs we recognize
            if dname.startswith( "hdl_" ):
                dir_path = os.path.join( root, dname )
                mtime = os.path.getmtime( dir_path )
                timestamp = ctime( mtime )
                logging.info( "removing dir: (created: %s) %s" % ( timestamp, dir_path ) )
                try:
                    shutil.rmtree( dir_path )
                except:
                    Nexcept += 1
                    type_, value, tb = sys.exc_info()
                    logging.error( "%s" % value )
        
        mtime = os.path.getmtime( root )
        timestamp = ctime( mtime )
        logging.info( "removing root: (created: %s) %s" % ( timestamp, root ) )
        try:
            shutil.rmtree( root )
        except:
            Nexcept += 1
            type_, value, tb = sys.exc_info()
            logging.error( "%s" % value )



def read_autoupdate( autoupdate_path ):
    update = False
    
    with open( autoupdate_path, 'r' ) as fd:
        text = fd.readline()
        text.strip()
        logging.info( "read_autoupdate() %s" % text )
        
        try:
            dt_dv = dateutil.parser.parse( text )       # e.g. 02-Jul-2019 in dataverse
            dt_au = datetime.now()                      # autoupdate
            if dt_dv > dt_au:
                logging.warn( "ignoring future date string" )
                return False
        except:
            logging.warn( "ignoring invalid date string" )
            return False
        
        logging.debug( "dt_dv: %s" % str( dt_dv ) )
        logging.debug( "dt_au: %s" % str( dt_au ) )
        
        secs_dv = mktime( dt_dv.timetuple() )
        secs_au = mktime( dt_au.timetuple() )
        seconds = secs_au - secs_dv
        
        seconds_per_day = 24 * 60 * 60
        
        logging.debug( "secs_dv: %d" % secs_dv )
        logging.debug( "secs_au: %d" % secs_au )
        logging.info( "seconds: %d" % seconds )
        logging.info( "seconds_per_day: %d" % seconds_per_day )
        
        if seconds < seconds_per_day:
            update = True
        else:
            logging.warn( "ignoring expired date string" )
    
    logging.info( "read_autoupdate() %s" % update )
    return update



def documents_by_handle( config_parser, handle_name, dst_dir, dv_format = "", check_ts = False ):
    global Nexcept
    
    to_csv = False
    logging.info( "documents_by_handle() check_ts: %s" % check_ts )
    logging.info( "documents_by_handle() handle_name: %s" % handle_name )
    logging.info( "dst_dir: %s, dv_format: %s" % ( dst_dir, dv_format ) )
    
    dv_host     = config_parser.get( "config", "dataverse_root" )
    dv_version  = config_parser.get( "config", "dv_version" )
    ristat_key  = config_parser.get( "config", "ristatkey" )
    ristat_name = config_parser.get( "config", "ristatname" )
    
    logging.info( "dv_host: %s" % dv_host )
    logging.info( "ristat_name: %s" % ristat_name )
    
    try:
        dv_connection = Connection( dv_host, ristat_key )
    except:
        logging.error( "Dataverse connection failed, with parameters:" )
        logging.info( "dv_host: %s" % dv_host )
        #logging.info( "ristat_key: %s" % ristat_key )
        type_, value, tb = sys.exc_info()
        logging.error( "etype: %s, value: %s" % ( type_, value ) )
        sys.exit( 1 )
    
    dataverse = dv_connection.get_dataverse( ristat_name )
    if not dataverse:
        logging.info( "ristat_key: %s" % ristat_key )
        logging.error( "COULD NOT GET A DATAVERSE CONNECTION" )
        sys.exit( 1 )
    
    logging.debug( "title: %s" % dataverse.title )
    #datasets = dataverse.get_datasets()
    
    #settings = DataFilter( '' )
    #papers = []
    ids = {}
    
    kwargs_xlsx2csv = { 
        "delimiter" : '|', 
        "lineterminator" : '\n'
        ,"quoting" : "csv.QUOTE_NONNUMERIC"
        #,"float_format" : "%.2f"
    }
    
    sep = str(u'|').encode('utf-8')
    kwargs_pandas = { 
        'sep' : sep, 
        'line_terminator' : '\n' 
    }
    

    tmp_dir = config_parser.get( "config", "tmppath" )
    dv_dir = "dataverse_src"
    download_dir = os.path.join( tmp_dir, dv_dir, dst_dir, handle_name )
    logging.info( "downloading dataverse files to: %s" % download_dir )
        
    if check_ts:
        # a place to save Autoupdate.txt
        if not os.path.exists( download_dir ):
            os.makedirs( download_dir )
    else:
        # download "normal" files
        if os.path.exists( download_dir ):
            empty_dir( download_dir )           # remove previous files
        if not os.path.exists( download_dir ):
            os.makedirs( download_dir )
    
    dv_items = dataverse.get_contents()
    nitems = len( dv_items )
    logging.info( "available dataverse handle items: %d" % nitems )
    
    for i, item in enumerate( dv_items ):
        logging.info( "handle item %d-of-%d" % ( i+1, nitems ) )
        # item dict keys: protocol, authority, persistentUrl, identifier, type, id
        handle = str( item[ 'protocol' ] ) + ':' + str( item[ 'authority' ] ) + "/" + str( item[ 'identifier' ] )
        logging.debug( "handle: %s" % handle )
        dv_handle = config_parser.get( "config", handle_name )
        logging.debug( "dv_handle: %s" % dv_handle )
        
        if handle != dv_handle:
            logging.info( "handle_name: %s, skipping handle: %s" % ( handle_name, handle ) )
        else:
            logging.info( "handle_name: %s, using handle: %s" % ( handle_name, handle ) )
            datasetid = item[ 'id' ]
            url  = "https://" + str( dv_host ) + "/api/datasets/" + str( datasetid )
            url += "?key=" + str( ristat_key )
            
            resp = {}
            try:
                resp = requests.get( url )
            except:
                logging.error( "FAILED TO GET DATA FROM URL: %s" % url )
                sys.exit( 1 )
            
            dataframe = resp.json()
            
            files = dataframe[ 'data' ][ 'latestVersion' ][ 'files' ]
            nfiles = len( files )
            logging.info( "number of files to download: %d" % nfiles )
            
            for nfile, dv_file in enumerate( files ):
                logging.info( "download %d-of-%d" % ( nfile+1, nfiles ) )
                logging.debug( "%s" % str( dv_file ))
                             
                datasetVersionId = str( dv_file[ "datasetVersionId" ] )
                version          = str( dv_file[ "version" ] )
                label            = str( dv_file[ "label" ] )
                
                if dv_version == "new":     # newer dataverse version
                    dataFile = dv_file[ "dataFile" ]
                    filename = str( dataFile[ 'filename' ] )
                else:                       # older dataverse version
                    dataFile = dv_file[ "datafile" ]
                    filename = str( dataFile[ 'name' ] )
                
                logging.info( "# of keys/vals in data file: %d" % len( dataFile ) )
                for key in dataFile:
                    logging.debug( "key: %s, val: %s" % ( key, dataFile[ key ] ) )
                
                paperitem = {}
                paperitem[ 'id' ]   = str( dataFile[ 'id' ] )
                originalFormatLabel = str( dataFile.get( 'originalFormatLabel', "" ) )
                # originalFormatLabel is no longer returned with new dataverse. 
                
                basename, ext = os.path.splitext( filename )
                
                logging.debug( "basename: %s, ext: %s, originalFormatLabel: %s" % ( basename, ext, originalFormatLabel ) )
                if dv_format == "original" and ext == ".tab" and originalFormatLabel in ["MS Excel (XLSX)", "MS Excel Spreadsheet"]:
                    filename = basename + ".xlsx"
                    logging.debug( "tab => xlsx: %s" % filename )
                
                paperitem[ 'name' ] = filename
                ids[ paperitem[ 'id'] ] = filename
                paperitem[ 'handle' ] = handle
                #paperitem[ 'url' ] = "http://data.sandbox.socialhistoryservices.org/service/download?id=%s" % paperitem[ 'id' ]
                
                url  = "https://%s/api/access/datafile/%s" % ( dv_host, paperitem[ 'id' ] )
                url += "?&key=%s&show_entity_ids=true&q=authorName:*" % str( ristat_key )
                if not dv_format == "":
                    url += "&format=original"
                logging.debug( url )
                
                #filename = paperitem[ 'name' ]
                
                logging.info( "filename: %s" % filename )
                
                filepath = os.path.join( download_dir, filename )
                logging.debug( "filepath: %s" % filepath )
                
                if check_ts:    # only check Autoupdate.txt
                    if filename == "Autoupdate.txt":
                        resp_in = requests.get( url )
                        with open( filepath, 'wb' ) as fd:
                            for chunk in resp_in.iter_content( chunk_size = 512 ):
                                fd.write( chunk )
                            os.fsync( fd )
                        
                        autoupdate = read_autoupdate( filepath )
                        return autoupdate
                
                else:
                    # read dataverse document from url, write contents to filepath
                    resp_in = requests.get( url )
                    with open( filepath, 'wb' ) as fd:
                        for chunk in resp_in.iter_content( chunk_size = 512 ):
                            fd.write( chunk )
                        os.fsync( fd )
                    
                    """
                    try:
                        if 'lang' in settings.datafilter:
                            varpat = r"(_%s)" % ( settings.datafilter[ 'lang' ] )
                            pattern = re.compile( varpat, re.IGNORECASE )
                            found = pattern.findall( paperitem[ 'name' ] )
                            if found:
                                papers.append( paperitem )
                        
                        if 'topic' in settings.datafilter:
                            varpat = r"(_%s_.+_\d+_+%s.|class|region)" % ( settings.datafilter[ 'topic' ], settings.datafilter[ 'lang' ] )
                            pattern = re.compile( varpat, re.IGNORECASE )
                            found = pattern.findall( paperitem[ 'name' ] )
                            if found:
                                papers.append( paperitem )
                        else:
                            if 'lang' not in settings.datafilter:
                                papers.append( paperitem )
                    except:
                        Nexcept += 1
                        type_, value, tb = sys.exc_info()
                        logging.error( "%s" % value )
                    """
    return ids



def documents_info( config_parser, language ):
    logging.debug( "documents_info()" )
    
    tmp_dir     = config_parser.get( "config", "tmppath" )
    dv_host     = config_parser.get( "config", "dataverse_root" )
    dv_version  = config_parser.get( "config", "dv_version" )
    api_root    = config_parser.get( "config", "api_root" )
    ristat_key  = config_parser.get( "config", "ristatkey" )
    ristat_name = config_parser.get( "config", "ristatname" )
    ristatdocs  = config_parser.get( "config", "hdl_documentation" )
    
    logging.info( "dv_host: %s" % dv_host )
    logging.debug( "ristat_key: %s" % ristat_key )
    logging.info( "ristat_name: %s" % ristat_name )
    
    download_dir = os.path.join( tmp_dir, "dataverse_src/doc" )
    download_fname = "doclist-" + language + ".json"
    download_path = os.path.join( download_dir, download_fname )
    
    logging.debug( "download_dir: %s" % download_dir )
    logging.debug( "download_fname: %s" % download_fname )
    logging.debug( "download_path: %s" % download_path )
    
    papers = []
    
    try:
        connection = Connection( dv_host, ristat_key )
        logging.debug( "connection succeeded" )
    except:
        logging.error( "connection failed" )
        #type_, value, tb = sys.exc_info()
        #logging.error( "%s" % value )
        etype = sys.exc_info()[ 0:1 ]
        value = sys.exc_info()[ 1:2 ]
        logging.error( "etype: %s, value: %s" % ( etype, value ) )
        return Response( json.dumps( papers ), mimetype = "application/json; charset=utf-8" )
    
    dataverse = connection.get_dataverse( ristat_name )
    if not dataverse:
        logging.info( "ristat_key: %s" % ristat_key )
        logging.error( "COULD NOT GET A DATAVERSE CONNECTION" )
        return Response( json.dumps( papers ), mimetype = "application/json; charset=utf-8" )
    
    logging.debug( "get_dataverse succeeded" )
    logging.debug( "title: %s" % dataverse.title )
    
    datatype = ""
    settings = DataFilter( { "lang" : language } )
    datafilter = settings.datafilter
    
    name_start = "ERRHS_" + str( datatype ) + "_"
    logging.debug( "name_start: %s" % name_start )
    
    for item in dataverse.get_contents():
        handle = str( item[ "protocol" ] ) + ':' + str( item[ "authority" ] ) + '/' + str( item[ "identifier" ] )
        if handle == ristatdocs:
            datasetid = item[ "id" ]
            url = "http://" + dv_host + "/api/datasets/" + str( datasetid ) + "/?&key=" + str( ristat_key )
            logging.debug( "url: %s" % url )
            
            resp = requests.get( url )
            dataframe = resp.json()
            
            for files in dataframe[ "data" ][ "latestVersion" ][ "files" ]:
                paperitem = {}
                
                if dv_version == "new":     # newer dataverse version
                    dataFile = files[ "dataFile" ]
                    filename = str( dataFile[ 'filename' ] )
                else:                       # newer dataverse version
                    dataFile = files[ "datafile" ]
                    filename = str( dataFile[ 'name' ] )
                
                paperitem[ "id" ] = str( dataFile[ "id" ] )
                paperitem[ "name" ] = filename
                paperitem[ "url" ] = "%s/service/download?id=%s" % ( api_root, paperitem[ "id" ] )
                logging.debug( "paperitem: %s" % paperitem )
                
                if datatype != "":      # use datatype to limit the returned documents
                    # find substring between the first two underscores
                    sub_name = ""
                    p1 = filename.find( "_" )
                    if p1 != -1:
                        p2 = filename.find( "_", p1+1 )
                        if p2 != -1:
                            sub_name = filename[ p1+1:p2 ]
                            logging.debug( "sub_name: %s" % sub_name )
                            try:
                                sub_digits = int( sub_name )
                                if sub_digits != datatype:
                                    continue    # datatype does not match: skip
                            except:
                                pass            # allow
                
                try:
                    if "lang" in settings.datafilter:
                        varpat = r"(_%s)" % ( settings.datafilter[ "lang" ] )
                        pattern = re.compile( varpat, re.IGNORECASE )
                        found = pattern.findall( paperitem[ "name" ] )
                        if found:
                            papers.append( paperitem )
                    else:   # paperitem without language specified: add
                        papers.append( paperitem )
                    
                    if "topic" in settings.datafilter:
                        varpat = r"(_%s_.+_\d+_+%s.|class|region)" % ( settings.datafilter[ "topic" ], settings.datafilter[ "lang" ] )
                        pattern = re.compile( varpat, re.IGNORECASE )
                        found = pattern.findall( paperitem[ "name" ] )
                        if found:
                            papers.append( paperitem )
                    else:
                        if "lang" not in settings.datafilter: 
                            papers.append( paperitem )
                except:
                    if "lang" not in settings.datafilter:
                        papers.append( paperitem )
    
    logging.debug( "papers in response:" )
    for paper in papers:
        logging.debug( paper )
    
    #return Response( json.dumps( papers ), mimetype = "application/json; charset=utf-8" )
    docs_json = json.dumps( papers )
    #logging.debug( type( docs_json ) )
    #logging.debug( docs_json )
    
    file_json = codecs.open( download_path, 'w', "utf-8" )
    file_json.write( docs_json.encode( "utf-8" ) )
    file_json.close()



def update_documentation( config_parser ):
    logging.info( "%s update_documentation()" % __file__ )
    
    handle_name = "hdl_documentation"
    logging.info( "retrieving documents from dataverse for handle name %s ..." % handle_name )
    dst_dir = "doc"
    dv_format = ""
    ids = documents_by_handle( config_parser, handle_name, dst_dir, dv_format )
    
    ndoc =  len( ids )
    logging.info( "%d documents retrieved from dataverse" % ndoc )
    if ndoc == 0:
        logging.info( "no documents, nothing downloaded." )
    
    for language in [ "ru", "en" ]:
        documents_info( config_parser, language )



def check_autoupdate( config_parser, dv_format ):
    logging.info( "%s check_autoupdate()" % __file__ )
    
    # Autoupdate.txt in vocab directory
    check_ts = True     # Read Autoupdate.txt only, no vocabs
    autoupdate = retrieve_vocabularies( config_parser, dv_format, check_ts )
    
    return autoupdate



def retrieve_vocabularies( config_parser, dv_format, check_ts = False ):
    global autoupdate
    
    logging.info( "%s retrieve_vocabularies() check_ts: %s" % ( __file__, check_ts ) )
    """
    if check_ts:
        return timestamp in Autoupdate.txt
    else:
        retrieve_vocabularies():
        -1- retrieves ERRHS_Vocabulary_*.tab files from dataverse
        -2- with copy_local = True stores them locally
        -3- stores the new data in MogoDB db = "vocabulary", collection = 'data'
    """
    
    # -1- Dataverse
    handle_name = "hdl_vocabularies"
    logging.info( "retrieving documents from dataverse for handle name %s ..." % handle_name )
    
    if dv_format == "original":
        dst_dir   = "vocab/xlsx"
        csv_dir   = "vocab/csv"
        ascii_dir = csv_dir
    else:
        dst_dir   = "vocab/tab"
        ascii_dir = dst_dir
    
    if check_ts: 
        autoupdate = documents_by_handle( config_parser, handle_name, dst_dir, dv_format, check_ts )
        return autoupdate
    else:
        dv_ids = documents_by_handle( config_parser, handle_name, dst_dir, dv_format )
    
    ndoc = len( dv_ids )
    logging.info( "%d documents retrieved from dataverse" % ndoc )
    if ndoc == 0:
        logging.info( "no documents, nothing to do." )
        return
    
    logging.debug( "keys in ids:" )
    for key in dv_ids:
        logging.debug( "key: %s, value: %s" % ( key, dv_ids[ key ] ) )

    return dv_ids



def copy_doc_src2dst():
    logging.info( "%s copy_doc_src2dst()" % __file__ )
    # Copy documentation from source (dataverse download) to destination directory (work dir)
    tmp_dir = config_parser.get( "config", "tmppath" )
    dv_dir_src = "dataverse_src"
    dv_dir_dst = "dataverse_dst"
    doc_dir = "doc"
    handle_name = "hdl_documentation"
    
    src_dir = os.path.join( tmp_dir, dv_dir_src, doc_dir, handle_name )
    dst_dir = os.path.join( tmp_dir, dv_dir_dst, doc_dir, handle_name )
    logging.info( "copying dataverse doc files from: %s" % src_dir )
    logging.info( "copying dataverse doc files to:   %s" % dst_dir )

    if os.path.exists( dst_dir ):
        empty_dir( dst_dir )        # remove previous files
    if not os.path.exists( dst_dir ):
        os.makedirs( dst_dir )
    
    dir_list = []
    dir_list = os.listdir( src_dir )
    dir_list.sort()
    
    for filename in dir_list:
        logging.info( "filename: %s" % filename )
        path_in  = os.path.join( src_dir, filename )
        path_out = os.path.join( dst_dir, filename )
        copy2( path_in, path_out )

    logging.info( "%d files copied" % len( dir_list ) )

    # copy ru & en listing files
    src_dir = os.path.join( tmp_dir, dv_dir_src, doc_dir )
    dst_dir = os.path.join( tmp_dir, dv_dir_dst, doc_dir )
    
    for filename in [ "doclist-en.json", "doclist-ru.json" ]:
        path_in  = os.path.join( src_dir, filename )
        path_out = os.path.join( dst_dir, filename )
        copy2( path_in, path_out )



def convert_vocabularies2csv( excel_package ):
    logging.info( "%s convert_vocabularies2csv()" % __file__ )
    
    tmp_dir = config_parser.get( "config", "tmppath" )
    vocab_dir = "vocab"
    handle_name = "hdl_vocabularies"
    
    xlsx_dir  = os.path.join( tmp_dir, "dataverse_src", vocab_dir, "xlsx", handle_name )
    csv_dir   = os.path.join( tmp_dir, "dataverse_dst", vocab_dir, "csv", handle_name )
    logging.info( "vocabulary  input: %s" % xlsx_dir )
    logging.info( "vocabulary output: %s" % csv_dir )
    
    if os.path.exists( csv_dir ):
        empty_dir( csv_dir )                # remove previous files
    if not os.path.exists( csv_dir ):
        os.makedirs( csv_dir )              # create destination dir
    
    dir_list = []
    if os.path.isdir( xlsx_dir ):
        dir_list = os.listdir( xlsx_dir )
        dir_list.sort()
    
        for xlsx_filename in dir_list:
            if xlsx_filename.endswith( ".xlsx" ):
                logging.info( "vocabulary filename: %s" % xlsx_filename )
                extra = ''  # vocab file contains both ru & en, so no "-ru" or "-en" extension
                
                if excel_package == "pandas":
                    xlsx2csv_pandas( xlsx_dir, xlsx_filename, csv_dir, extra )
                elif excel_package == "openpyxl":
                    xlsx2csv_openpyxl( xlsx_dir, xlsx_filename, csv_dir, extra )
                elif excel_package == "tablib":
                    xlsx2csv_tablib( xlsx_dir, xlsx_filename, csv_dir, extra )
                else:
                    logging.error( "excel_package: %s not supported" % excel_package )
                    logging.error( "EXIT" )
                    print( "EXIT" )
                    sys.exit( 1 )
            else:
                logging.info( "skip filename: %s" % xlsx_filename )
                continue



def merge_vocabs( vocab_csv_dir ):
    logging.info( "%s merge_vocab()" % __file__ )
    
    lexicon = []
    len_totvocab = 0
    
    dir_list = []
    if os.path.isdir( vocab_csv_dir ):
        dir_list = os.listdir( vocab_csv_dir )
    
    logging.info( "using csv directory: %s" % vocab_csv_dir )
    
    for i, filename in enumerate( dir_list ):
        basename, ext = os.path.splitext( filename )
        if ext != ".csv":
            logging.info( "skipping: %s" % filename )
            continue
        
        logging.info( "filename %s" % filename )
        pathname = os.path.join( vocab_csv_dir, filename )
        
        with open( pathname, 'r') as f:
            data = f.read()
            csvio = StringIO( str( data ) )
            dataframe = pd.read_csv( csvio, sep = '|', dtype = 'unicode' )
    
        # fetch columns from dataverse vocabulary file
        filter_columns = []
        mapping = {}
        for col in dataframe.columns:
            findval = re.search( r'RUS|EN|ID|TOPIC_ID|TOPIC_ROOT|DATATYPE|YEAR|basisyear', col )
            if findval:
                mapping[ col ] = findval.group( 0 )
                filter_columns.append( col )
        
        vocab = {}
        if filter_columns:
            logging.info( "filter_columns: %s" % filter_columns )
            vocab = dataframe[ filter_columns ]
            newcolumns = []
            for field in vocab:
                value = mapping[ field ]
                newcolumns.append( value )
            
            vocab.columns = newcolumns
            vocab = vocab.dropna()
            vocab[ 'vocabulary' ] = basename
            len_vocab = len( vocab )
            len_totvocab += len_vocab
            lexicon.append( vocab )
            
            logging.info( "id: %s, filename: %s, items: %d" % ( i, filename, len_vocab ) )
        else:
            logging.warning( "No filter_columns" )
    
    logging.info( "lexicon contains %d vocabularies containing %d items in total" % ( len( lexicon ), len_totvocab ) )
    
    # concatenate the vocabularies with pandas
    return pd.concat( lexicon, sort = True )



def mongo_store_vocabularies():
    logging.info( "%s mongo_store_vocabularies()" % __file__ )
    
    # Going to up update MongoDB contents; clear it first
    clear_mongo( mongo_client )
    
    # parameters to retrieve the vocabulary files
    dv_host = config_parser.get( "config", "dataverse_root" )
    apikey  = config_parser.get( "config", "ristatkey" )
    dbname  = config_parser.get( "config", "vocabulary" )
    
    logging.debug( "dv_host:   %s" % dv_host )
    logging.debug( "apikey: %s" % apikey )
    logging.debug( "dbname: %s" % dbname )
    
    vocab_json = [ {} ]
    tmp_dir = config_parser.get( "config", "tmppath" )
    handle_name = "hdl_vocabularies"
    abs_ascii_dir = os.path.join( tmp_dir, "dataverse_dst", "vocab/csv", handle_name )
    
    # do not read again from dataverse
    big_vocabulary = merge_vocabs( abs_ascii_dir  )
    
    #print big_vocabulary.to_json( orient = 'records' )
    vocab_json = json.loads( big_vocabulary.to_json( orient = 'records' ) )  # type: <type 'list'>
    
    """
    vocab_json0 = vocab_json[ 0 ]
    for key in vocab_json0:
        logging.info( "key: %s, value: %s" % ( key, vocab_json0[ key ] ) )
    
    # there are 7 keys per dictionary entry, e.g vocab_json0:
    key: basisyear,  value: None
    key: EN,         value: Male
    key: vocabulary, value: ERRHS_Vocabulary_modclasses
    key: DATATYPE,   value: None
    key: YEAR,       value: None
    key: RUS,        value: мужчины
    key: ID,         value: MOD_1.01_1
    """
    
    logging.info( "processing %d vocabulary items..." % len( vocab_json ) )
    for item in vocab_json:
        #logging.debug( item )
        if 'YEAR' in item:
            item[ 'YEAR' ] = re.sub( r'\.0', '', str( item[ 'YEAR' ] ) )
            #logging.debug( "YEAR: %s" % item[ 'YEAR' ] )
        if 'basisyear' in item:
            item[ 'basisyear' ] = re.sub( r'\.0', '', str( item[ 'basisyear' ] ) )
            #logging.debug( "basisyear: %s" % item[ 'basisyear' ] )
    
    dbname_vocab = config_parser.get( "config", "vocabulary" )
    db_vocab = mongo_client.get_database( dbname_vocab )
    logging.info( "inserting vocabulary in mongodb '%s'" % dbname_vocab )
    result = db_vocab.data.insert( vocab_json )



def retrieve_handle_docs( config_parser, handle_name, dv_format = "" ):
    logging.info( "retrieve_handle_docs()" )

    logging.info( "retrieving documents from dataverse for handle name %s ..." % handle_name )
    dst_dir = "xlsx"
    ids = documents_by_handle( config_parser, handle_name, dst_dir, dv_format )
    ndoc =  len( ids )
    if ndoc == 0:
        logging.info( "no documents retrieved." )
        return
    else:
        logging.info( "%d documents for handle %s retrieved from dataverse" % ( ndoc, handle_name ) )
    
    logging.debug( "keys in ids:" )
    for key in ids:
        logging.debug( "key: %s, value: %s" % ( key, ids[ key ] ) )



def row_count( config_parser, language ):
    logging.debug( "row_count()" )

    configpath = RUSSIANREPO_CONFIG_PATH
    if not os.path.isfile( configpath ):
        print( "in %s" % __file__ )
        print( "configpath %s FILE DOES NOT EXIST" % configpath )
        print( "EXIT" )
        sys.exit( 1 )
    
    logging.debug( "using configparser: %s" % configpath )

    config_parser.read( configpath )
    
    dbtable_name = "dbtable" + '_' + language
    dbtable  = config_parser.get( "config", dbtable_name )
    dbhost   = config_parser.get( "config", "dbhost" )
    dbname   = config_parser.get( "config", "dbname" )
    user     = config_parser.get( "config", "dblogin" )
    password = config_parser.get( "config", "dbpassword" )
    
    connection_string = "host = '%s' dbname = '%s' user = '%s' password = '%s'" % ( dbhost, dbname, user, password )
    logging.debug( "connection_string: %s" % connection_string )
    
    try:
        pg_connection = psycopg2.connect( connection_string )
    except:
        etype = sys.exc_info()[ 0:1 ]
        value = sys.exc_info()[ 1:2 ]
        logging.error( "row_count() %s, %s\n" % ( etype, value ) )
        raise RuntimeError()
    
    cursor = pg_connection.cursor()
    sql = "SELECT COUNT(*) FROM %s;" % dbtable
    logging.info( sql )
    
    try:
        cursor.execute( sql )
        data = cursor.fetchall()
        count = data[0][0]
        logging.info( "row count: %d" % count )
        
        pg_connection.commit()
        cursor.close()
        pg_connection.close()
    except:
        logging.error( "row_count() failed:" )
        type_, value, tb = sys.exc_info()
        logging.error( "%s" % value )



def clear_postgres_table( config_parser, language ):
    logging.info( "clear_postgres()" )

    configpath = RUSSIANREPO_CONFIG_PATH
    if not os.path.isfile( configpath ):
        print( "in %s" % __file__ )
        print( "configpath %s FILE DOES NOT EXIST" % configpath )
        print( "EXIT" )
        sys.exit( 1 )
    
    logging.info( "using configparser: %s" % configpath )

    config_parser.read( configpath )
    
    dbtable_name = "dbtable" + '_' + language
    dbtable  = config_parser.get( "config", dbtable_name )
    dbhost   = config_parser.get( "config", "dbhost" )
    dbname   = config_parser.get( "config", "dbname" )
    user     = config_parser.get( "config", "dblogin" )
    password = config_parser.get( "config", "dbpassword" )
    
    connection_string = "host = '%s' dbname = '%s' user = '%s' password = '%s'" % ( dbhost, dbname, user, password )
    logging.info( "connection_string: %s" % connection_string )

    try:
        pg_connection = psycopg2.connect( connection_string )
    except:
        etype = sys.exc_info()[ 0:1 ]
        value = sys.exc_info()[ 1:2 ]
        logging.error( "clear_postgres_table() %s, %s\n" % ( etype, value ) )
        raise RuntimeError()

    sql = "TRUNCATE TABLE %s;" % dbtable
    logging.info( sql )
    cursor = pg_connection.cursor()
    cursor.execute( sql )
    
    pg_connection.commit()
    cursor.close()
    pg_connection.close()



def store_handle_docs( config_parser, handle_name, language ):
    logging.info( "" )
    logging.info( "store_handle_docs() %s" % handle_name )
    
    tmp_dir = config_parser.get( "config", "tmppath" )
    csv_dir_l = "csv-" + language
    csv_dir  = os.path.join( tmp_dir, "dataverse_dst", csv_dir_l, handle_name )
    dir_list = []
    if os.path.isdir( csv_dir ):
        dir_list = os.listdir( csv_dir )
    
    logging.info( "using csv directory: %s" % csv_dir )

    configpath = RUSSIANREPO_CONFIG_PATH
    if not os.path.isfile( configpath ):
        print( "in %s" % __file__ )
        print( "configpath %s FILE DOES NOT EXIST" % configpath )
        print( "EXIT" )
        sys.exit( 1 )
    
    
    logging.info( "using configparser: %s" % configpath )

    config_parser.read( configpath )
    
    dbtable_name = "dbtable" + '_' + language
    dbtable  = config_parser.get( "config", dbtable_name )
    dbhost   = config_parser.get( "config", "dbhost" )
    dbname   = config_parser.get( "config", "dbname" )
    user     = config_parser.get( "config", "dblogin" )
    password = config_parser.get( "config", "dbpassword" )
    
    connection_string = "host = '%s' dbname = '%s' user = '%s' password = '%s'" % ( dbhost, dbname, user, password )
    logging.info( "connection_string: %s" % connection_string )

    try:
        pg_connection = psycopg2.connect( connection_string )
    except:
        etype = sys.exc_info()[ 0:1 ]
        value = sys.exc_info()[ 1:2 ]
        logging.error( "store_handle_docs() %s, %s\n" % ( etype, value ) )
        raise RuntimeError()
    
    cursor = pg_connection.cursor()

    for filename in dir_list:
        root, ext = os.path.splitext( filename )
        if root.startswith( "ERRHS_" ) and ext == ".csv":
            logging.info( "use: %s, to table: %s" % ( filename, dbtable ) )
            in_pathname = os.path.abspath( os.path.join( csv_dir, filename ) )
            logging.debug( in_pathname )
            #test_csv_file( pathname )
            
            #out_pathname = write_psv_file( csv_dir, filename )
            #psv_file = codecs.open( out_pathname, 'r', encoding = "utf-8" )
            #cursor.copy_from( psv_file, dbtable, sep ='|' )
            
            stringio_file = filter_csv( config_parser, csv_dir, filename )
            cursor.copy_from( stringio_file, dbtable, sep = '|' )
            #cursor.copy_from( stringio_file, dbtable, sep = '|', null = "None" )
            
            #csv_strings.close()  # close object and discard memory buffer
            #csvfile.close()
            
            # debug strange record duplications
            pg_connection.commit()
            row_count( config_parser, language )
            
        else:
            logging.info( "skip: %s" % filename )

        #print( "break" )
        #break
    
    ndoc = len( dir_list )
    logging.info( "%d documents for handle %s stored in table %s" % ( ndoc, handle_name, dbtable ) )
    
    pg_connection.commit()
    cursor.close()
    pg_connection.close()



def test_csv_file( path_name ):
    csv_file = open( path_name, 'r' )
    #csv_file = codecs.open( path_name, 'r', encoding = 'utf-8' )
    nlines = 0
    
    for line in csv_file:
        cnt = line.count( '|' )
        fields = line.split( '|' )
        nfields = len( fields )
        print( "%d: %d" % ( nline, nfields ) )
        nlines += 1
        
    print( "%d" % nlines )



def filter_csv( config_parser, csv_dir, in_filename ):
    global Nexcept
    global pkey
    
    logging.info( "filter_csv() %s" % in_filename )
    
    # Notice: the applied filtering is reflected in the returned out_file, 
    # which is copied to a postgres table; the input csv file is _unchanged_.
    
    """
    # FL-10-Dec-2019 Assume tablib does not have the pandas problem
    # rounding of data in value column is specified in ERRHS_Vocabulary_units.csv
    # Equality of pos_rus and pos_eng is a hack, used to flag decimals from either rus or eng
    vocab_units = dict()
    if in_filename.endswith( "-ru.csv" ):
        vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 0, 0 )
    elif in_filename.endswith( "-en.csv" ):
        vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 1, 1 )
    else:
        logging.error( "in_filename does not end with either '-ru.csv' or '-en.csv'" )
    #logging.info( "vocab_units: %s" % str( vocab_units ) )
    """
    
    # dataverse column names
    dv_column_names = [
        "id", 
        "territory", 
        "ter_code", 
        "town", 
        "district", 
        "year", 
        "month", 
        "value", 
        "value_unit", 
        "value_label", 
        "datatype", 
        "histclass1", 
        "histclass2", 
        "histclass3", 
        "histclass4", 
        "histclass5", 
        "histclass6", 
        "histclass7", 
        "histclass8", 
        "histclass9", 
        "histclass10", 
        "class1", 
        "class2", 
        "class3", 
        "class4", 
        "class5", 
        "class6", 
        "class7", 
        "class8", 
        "class9", 
        "class10", 
        "comment_source", 
        "source", 
        "volume", 
        "page", 
        "naborshik_id", 
        "comment_naborshik", 
        "base_year"
    ]
    
    # some extra columns in postgres table
    pg_column_names = list( dv_column_names )
    # FL-13-Mar-2017: dropped column "indicator_id"
    #pg_column_names.insert( 0, "indicator_id" ) # pre-pended for postgres table
    pg_column_names.append( "indicator" )       # appended for for postgres table
    pg_column_names.append( "valuemark" )       # appended for for postgres table
    pg_column_names.append( "timestamp" )       # appended for for postgres table
    # FL-13-Mar-2017: added primary key (after "timestamp")
    pg_column_names.append( "pk" )              # appended for for postgres table
    
    ncolumns_dv = len( dv_column_names )
    logging.debug( "# of dv_columns: %d" % ncolumns_dv )
    
    in_pathname = os.path.abspath( os.path.join( csv_dir, in_filename ) )
    csv_file = open( in_pathname, 'r' )
    #csv_file = codecs.open( in_pathname, 'r', encoding = 'utf-8' )
    
    """
    root, ext = os.path.splitext( in_filename )
    out_pathname = os.path.abspath( os.path.join( csv_dir, root + ".psv" ) )
    print( out_pathname )
    out_file = codecs.open( out_pathname, 'w', encoding = 'utf-8' )
    """
    
    out_file = StringIO()               # in-memory file
    
    nline = 0
    nskipped = 0
    comment_length_max = 0
    
    nfiltered = 0                       # count filtered values reduce decimals)
    nstripped = 0                       # count stripped fields
    stripped_fields  = set()            # collect only unique fields
    affected_headers = set()            # corresponding headers
    
    for line in csv_file:
        nline += 1
        #logging.info( "line %d: %s" % ( nline, line ) )
        line = line.strip( '\n' )       # remove trailing \n
        #logging.info( "%d in: %s" % ( nline, line ) )
        #print( "# new lines: %d" % line.count( '\n' ) )
        
        if len( line ) == line.count( '|' ):    # only separators, no data
            nskipped += 1
            continue
        
        fields = line.split( '|' )
        if nline == 1:
            line_header = line
            nfields_header = len( fields )              # nfields of header line
            
            # need a list as result, otherwise error in comparison below: 
            # TypeError: 'itertools.imap' object has no attribute '__getitem__'
            csv_header_names = list( map( str.lower, fields ) )
            logging.debug( "# of csv header fields: %d" % nfields_header )
            ndiff = nfields_header - ncolumns_dv
            #logging.info( "ndiff: %d" % ndiff )
            
            # header names must match predefined dv_column_names
            # (but after base_year there may be several empty columns in the input)
            skip_file = False
            if len( dv_column_names ) > len( csv_header_names ):
                msg = "skipping bad file %s" % in_filename
                logging.warning( msg )
                #print( msg )
                msg = "wrong header structure: \n%s" % line_header
                logging.warning( msg )
                #print( msg )
                skip_file = True
                break
            for i in range( ncolumns_dv ):
                if dv_column_names[ i ] != csv_header_names[ i ]:
                    msg = "skipping bad file %s" % in_filename
                    logging.warning( msg )
                    #print( msg )
                    msg = "wrong header structure: \n%s" % line_header
                    logging.warning( msg )
                    #print( msg )
                    skip_file = True
                    break
            if skip_file:
                break
                
            continue        # do not store header line
        else:
            # Keep the bridging '.' but replace trailing '.' by ". " (add a space)
            # So the (unordered) '. ' are easily recognizable in the RiStat requests
            nfields = len( fields )
            if nfields != nfields_header:
                msg = "skipping bad data line # %d" % nline
                logging.warning( msg )
                #print( msg )
                msg = "# of data fields (%d) does not match # of header fields (%d)" % ( nfields, nfields_header )
                logging.warning( msg )
                #print( msg )
                msg = "header: %s" % line_header
                logging.warning( msg )
                #print( msg )
                msg = "data: %s" % line
                logging.warning( msg )
                #print( msg )
                continue
            
            # strip leading and trailing white space, 
            # but only for the fields that are used in translations
            units       = [ "VALUE_UNIT" ]
            histclasses = [ "HISTCLASS1", "HISTCLASS2", "HISTCLASS3", "HISTCLASS4", "HISTCLASS5",  "HISTCLASS6", "HISTCLASS7", "HISTCLASS8", "HISTCLASS9", "HISTCLASS10", ]
            modclasses  = [ "CLASS1", "CLASS2", "CLASS3", "CLASS4", "CLASS5",  "CLASS6", "CLASS7", "CLASS8", "CLASS9", "CLASS10", ]
            translate_list = units + histclasses + modclasses
            
            for i in range( nfields ):
                csv_header_name = csv_header_names[ i ]
                if csv_header_name in translate_list:
                    in_field = fields[ i ]
                    field_stripped = in_field.strip()
                    if in_field != field_stripped:
                        nstripped += 1
                        affected_headers.add( csv_header_name )
                        stripped_fields.add( in_field )
                        fields[ i ] = field_stripped
                        #msg = "stripped: |%s| -> |%s|" % ( field_in, fields[ i ] )
                        #logging.warning( msg ); print( msg )
            
            nzaphc = 0
            for i in reversed( range( nfields ) ):      # histclass fields
                #print( "%2d %s: %s" % ( i, csv_header_names[ i ], fields[ i ] ) )
                if csv_header_names[ i ].startswith( "histclass" ):     # historical
                    if fields[ i ] == ".":
                        fields[ i ] = ". "
                        nzaphc += 1
                    else:
                        break
            
            nzapc = 0
            for i in reversed( range( nfields ) ):  # class fields
                #print( "%2d %s: %s" % ( i, csv_header_names[ i ], fields[ i ] ) )
                if csv_header_names[ i ].startswith( "class" ):         # modern
                    if fields[ i ] == ".":
                        fields[ i ] = ". "
                        nzapc += 1
                    else:
                        break
            
            """
            # replace the value dot fields by an empty string, 
            # because the dots hamper COUNT, and we want them in the response.
            for i in range( nfields ):
                #print( "%2d %s: %s" % ( i, csv_header_names[ i ], fields[ i ] ) )
                if csv_header_names[ i ] == "value" :
                    if fields[ i ] == ".":
                        fields[ i ] = "None"
            """
            
            # FL-06-Mar-2019 integer strings became float strings ending with .0
            # should be solved when writing csv files, but pandas did not do what i want!
            #if in_filename.startswith( "ERRHS_1_02_data_1897" ):
            # FL-14-May-2019 also the other years
            """
            # FL-10-Dec-2019 Assume tablib does not have the pandas problem
            if in_filename.startswith( "ERRHS_1_02_data_" ):
                # check some [hist]class values
                histclass1_pos = csv_header_names.index( "histclass1" )
                in_field = fields[ histclass1_pos ]
                try: 
                    in_field_float = float( in_field )
                    if str( in_field ).endswith( '.0' ): 
                        decimals = 0
                        in_field_round = round( in_field_float, decimals )
                        in_field_new = str( long( round( in_field_float, decimals ) ) )
                        fields[ histclass1_pos ] = in_field_new
                except:
                    pass
                
                class2_pos = csv_header_names.index( "class2" )
                in_field = fields[ class2_pos ]
                try: 
                    in_field_float = float( in_field )
                    if str( in_field ).endswith( '.0' ): 
                        decimals = 0
                        in_field_round = round( in_field_float, decimals )
                        in_field_new = str( long( round( in_field_float, decimals ) ) )
                        fields[ class_pos ] = in_field_new
                except:
                    pass
            """
            
            # check comment_source length
            comment_pos = csv_header_names.index( "comment_source" )
            comment = fields[ comment_pos ]
            comment_length = len( comment )
            comment_length_max = max( comment_length_max, comment_length )
            if comment_length > COMMENT_LENGTH_MAX_DB:
                fields[ comment_pos ] = ""      # because it is unicode we cannot just chop it
                msg = "too long comment in line:"
                logging.warning( msg )
                #print( msg )
                logging.warning( line )
                #print( line )
            
            """
            # FL-10-Dec-2019 Assume tablib does not have this problem
            # check missing datatype
            datatype_pos = csv_header_names.index( "datatype" )
            datatype = fields[ datatype_pos ]
            if len( datatype ) == 0 or datatype == '.':
                msg = "missing datatype in line:"
                logging.warning( msg )
                #print( msg )
                logging.warning( line )
                #print( line )
            else:   # chop spurious decimals of stupid spreadsheets
                fields[ datatype_pos ] = "%4.2f" % float( datatype )
            """
            
            """
            # FL-10-Dec-2019 Value rounding done when creating csv files
            # round value to specified number of decimals
            value_pos = csv_header_names.index( "value" )
            value_str = fields[ value_pos ]
            value_unit_pos = csv_header_names.index( "value_unit" )
            value_unit = fields[ value_unit_pos ]
            
            decimals = 0
            try:
                decimals = int( vocab_units[ value_unit ] )
                logging.debug( "value: %s, value_unit: %s, decimals: %s" % ( value_str, value_unit, decimals ) )
            except:
                decimals = 0
                pass
                
            if value_str.isdigit():
                pass    # all digits, no change
            else:
                try:
                    value_float = float( value_str )
                    value_round = round( value_float, decimals )
                    if decimals == 0:
                        value_new = str( long( round( value_float, decimals ) ) )
                    else:
                        value_new = str( round( value_float, decimals ) )
                    
                    if value_new != value_str:
                        nfiltered += 1
                        fields[ value_pos ] = value_new     # replace
                except:
                    pass    # no change
            """
        
        #print( "|".join( fields ) )
        if ndiff > 0:
            npop = ndiff
            for _ in range( npop ):
                fields.pop()            # skip trailing n fields
            #logging.debug( "# of fields popped: %d" % npop )
        elif ndiff < 0:
            napp = abs( ndiff ) - 1
            for _ in range( napp ):
                fields.append( "" )
            #logging.debug( "# of fields added: %d" % napp )
        #print( "# of fields: %d" % len( fields ) )
        
        """
        # FL-10-Dec-2019 Assume tablib does not have this problem
        # base_year, must be integer
        base_year_idx = None
        try:
            base_year_idx = csv_header_names.index( "base_year" )   # 38, 37, ...?
            try:
                base_year_in  = fields[ base_year_idx ]             # e.g. "1897.0" due to unwanted pandas xlsx input conversion?
                base_year_out = str( int( float( fields[ base_year_idx ] ) ) )
                if base_year_in != base_year_out:
                    fields[ base_year_idx ] = base_year_out
            except:
                logging.warning( "base_year not integer: %s" % fields[ base_year_idx ] )
                try:
                    fields[ base_year_idx ] = "0"
                except:
                    Nexcept += 1
                    logging.info( "%d: in: %s" % ( nline, line ) )
                    logging.info( "%d out: %s" % ( nline, "|".join( fields ) ) )
                    type_, value, tb = sys.exc_info()
                    msg = "%s: %s" % ( type_, value )
                    logging.error( msg )
                    #sys.stderr.write( "%s\n" % msg )
        except ValueError:
            pass
        """
        
        #print( 1, "|".join( fields ) )
        fields.append( "0" )                # indicator field, not in csv file
        #print( 2, "|".join( fields ) )
        fields.append( "false" )            # valuemark field, not in csv file
        #print( 3, "|".join( fields ) )
        
        """
        # valuemark must be true or false, but is not in the input data
        valuemark_idx = base_year_idx + 2
        if fields[ valuemark_idx ] not in ( "true", "false" ):
            fields[ valuemark_idx ] = "false"
        """
        
        fields.append( "now()" )            # timestamp field, not in csv file
        
        # column indicator_id should become the primairy key
        # FL-13-Mar-2017: added primary key after "timestamp"
        if pkey is None: 
            pkey = 1
        fields.append( str( pkey ) )        # global
        pkey += 1
        
        # FL-13-Mar-2017: dropped column "indicator_id"
        #fields.insert( 0, "0" )             # prepend indicator_id field, not in csv file
        
        #print( "|".join( fields ) )
        
        table_line = "|".join( fields )
        logging.debug( "fields in table record: %d" % len( fields ))
        """
        if nline == 2:
            print( "%d fields" % len( fields ) )
            print( short_line )
        
            for i in range( len( fields ) ):
                print( "%2d %s: %s" % ( i, dv_column_names[ i ], fields[ i ] ) )
        """
        #logging.info( "%d: in: %s" % ( nline, line ) )
        #logging.info( "%d out: %s" % ( nline, table_line ) )
        
        out_file.write( "%s\n" % table_line )
    
    out_file.seek( 0 )   # start of the stream
    #out_file.close()    # closed by caller!: closing discards memory buffer
    csv_file.close()
    
    logging.info( "lines written to csv file: %d (including header line)" % (nline - nskipped) )
    
    if comment_length_max > COMMENT_LENGTH_MAX_DB:
        logging.info( "WARNING: comment_length_max: %d, length available %d" % ( comment_length_max, COMMENT_LENGTH_MAX_DB ) )
    
    if nskipped != 0:
        logging.info( "empty lines (|-only) skipped: %d" % nskipped )
    
    if nstripped != 0:
        logging.info( "%d fields stripped, of which unique: %d" % ( nstripped, len( stripped_fields ) ) )
        logging.info( "affected_headers: %s" % affected_headers )
        for field in stripped_fields:
            logging.info( "leading/trailing whitespace: |%s|" % field )
    
    if nfiltered != 0:
        logging.info( "%d value fields were filtered to reduce the # of decimals" % nfiltered )
    
    #return out_pathname
    return out_file



def update_handle_docs( config_parser, mongo_client, language ):
    logging.info( "" )
    logging.info( "update_handle_docs() language = %s" % language )
    
    configpath = RUSSIANREPO_CONFIG_PATH
    logging.info( "using configparser: %s" % configpath )
    # classupdate() uses postgresql access parameters from cpath contents
    # fetching historic and modern class data from postgresql table, 
    # either from the 'ru' or from the 'en' table
    classdata = classupdate( configpath, language )
    
    # for the class data, use language dependent mongo collections
    dbname = config_parser.get( "config", "vocabulary" )
    dbname += ( '_' + language )
    logging.info( "inserting historic and modern class data in mongodb '%s'" % dbname )
    
    db = mongo_client.get_database( dbname )
    result = db.data.insert( classdata )



def clear_mongo( mongo_client ):
    logging.info( "clear_mongo()" )
    
    dbname_vocab = config_parser.get( "config", "vocabulary" )
    logging.info( "delete all documents from collection 'data' in mongodb db '%s'" % dbname_vocab )
    db_vocab = mongo_client.get_database( dbname_vocab )
    # drop the documents from collection 'data'; same as: db.drop_collection( coll_name )
    db_vocab.data.drop()
    
    dbname_vocab_ru = dbname_vocab + "_ru"
    logging.info( "delete all documents from collection 'data' in mongodb db '%s'" % dbname_vocab_ru )
    db_vocab = mongo_client.get_database( dbname_vocab_ru )
    db_vocab.data.drop()
    
    dbname_vocab_en = dbname_vocab + "_en"
    logging.info( "delete all documents from collection 'data' in mongodb db '%s'" % dbname_vocab_en )
    db_vocab = mongo_client.get_database( dbname_vocab_en )
    db_vocab.data.drop()
    
    logging.info( "clearing mongodb cache" )
    db_cache = mongo_client.get_database( 'datacache' )
    db_cache.data.drop()      # remove the collection 'data'



def topic_counts( config_parser, langage ):
    logging.info( "topic_counts()" )
    
    configpath = RUSSIANREPO_CONFIG_PATH
    if not os.path.isfile( configpath ):
        logging.error( "in %s" % __file__ )
        logging.error( "configpath %s FILE DOES NOT EXIST" % configpath )
        logging.error( "EXIT" )
        sys.exit( 1 )
    
    logging.info( "using configparser: %s" % configpath )

    config_parser.read( configpath )
    
    dbtable_name = "dbtable" + '_' + language
    dbtable  = config_parser.get( "config", dbtable_name )
    dbhost   = config_parser.get( "config", "dbhost" )
    dbname   = config_parser.get( "config", "dbname" )
    user     = config_parser.get( "config", "dblogin" )
    password = config_parser.get( "config", "dbpassword" )
    
    connection_string = "host = '%s' dbname = '%s' user = '%s' password = '%s'" % ( dbhost, dbname, user, password )
    logging.info( "connection_string: %s" % connection_string )

    try:
        pg_connection = psycopg2.connect( connection_string )
    except:
        etype = sys.exc_info()[ 0:1 ]
        value = sys.exc_info()[ 1:2 ]
        logging.error( "topic_counts() %s, %s\n" % ( etype, value ) )
        raise RuntimeError()
    
    cursor = pg_connection.cursor( cursor_factory = psycopg2.extras.NamedTupleCursor )

    sql_topics  = "SELECT datatype, topic_name FROM datasets.topics"
    sql_topics += " ORDER BY datatype"
    logging.info( sql_topics )
    cursor.execute( sql_topics )
    resp = cursor.fetchall()
    
    main_topics = [ "1", "2", "3", "4", "5", "6", "7" ]
    for record in resp:
        datatype   = record.datatype
        topic_name = record.topic_name
        if datatype not in main_topics:
            #print( datatype, topic_name )
            sql_count  = "SELECT base_year, COUNT(*) AS count FROM russianrepository"
            sql_count += " WHERE datatype = '%s'" % datatype
            sql_count += " GROUP BY base_year ORDER BY base_year"
            logging.debug( sql_count )
            
            cursor.execute( sql_count )
            cnt_resp = cursor.fetchall()
            cnt_dict = {}
            for cnt_rec in cnt_resp:
                #print( cnt_rec )
                cnt_dict[ cnt_rec.base_year ] = int( cnt_rec.count )    # strip trailing 'L'
            #print( cnt_dict )
            logging.info( "datatype: %s , topic_name: %s, counts: %s" % ( datatype, topic_name, str( cnt_dict ) ) )
        else:
            #print( "skip main topic :", datatype, topic_name )
            pass
    
    pg_connection.commit()
    cursor.close()
    pg_connection.close()



def load_vocab( config_parser, vocab_fname, vocab, pos_rus, pos_eng ):
    global Nexcept
    
    logging.info( "load_vocab() vocab_fname: %s" % vocab_fname )
    # if pos_extar is not None, it is needed to make the keys and/or values unique
    handle_name = "hdl_vocabularies"
    tmp_dir = config_parser.get( "config", "tmppath" )
    vocab_dir = os.path.join( tmp_dir, "dataverse_dst", "vocab/csv", handle_name )
    vocab_path = os.path.join( vocab_dir, vocab_fname )
    logging.debug( "vocab_dir: %s" % vocab_dir )
    logging.debug( "vocab_path: %s" % vocab_path )
    
    #vocab_file = open( vocab_path, "r" )
    vocab_file = codecs.open( vocab_path, "r", encoding = 'utf-8' )
    
    nline = 0
    for csv_line in iter( vocab_file ):
        csv_line.strip()        # zap '\n'
        logging.debug( csv_line )
        if nline == 0:
            pass        # skip header
        else:
            parts_list = csv_line.split( '|' )
            parts_dict = dict( enumerate( parts_list ) )
            
            rus = parts_dict.get( pos_rus, "" ).strip()
            eng = parts_dict.get( pos_eng, "" ).strip()
            
            if vocab_fname == "ERRHS_Vocabulary_regions.csv":
                # The regions (territorium) vocab is special: 
                # vocab = dict(), not bidict()
                # the original rus terms contain a lot of noise, but the codes 
                # and eng translations are unique and OK. For both forward and 
                # inverse lookup the code is the key, and either rus or eng the value. 
                terr = parts_dict.get( 0, "" ).strip()
                terr_d = { "terr" : terr }
                terr_s = json.dumps( terr_d )
                rus_eng_d = { "rus" : rus, "eng" : eng }
                logging.debug( "terr: %s, rus_eng: %s %s" % ( terr, rus_eng_d[ "rus" ], rus_eng_d[ "eng" ] ) )
                rus_eng_s = json.dumps( rus_eng_d )
                try:
                    vocab[ terr_s ] = rus_eng_s
                except:
                    Nexcept += 1
                    type_, value, tb = sys.exc_info()
                    msg = "%s: %s %s" % ( type_, vocab_fname, value )
                    logging.error( msg )
                    #sys.stderr.write( "%s\n" % msg )
            elif vocab_fname == "ERRHS_Vocabulary_units.csv":
                if pos_rus == pos_eng:              # a bit of a hack
                    # vocab is a dict(), not bidict(), 
                    # used for decimals, either from rus or from eng
                    decimals_str = parts_dict.get( 2, "0" )
                    decimals = int( float( decimals_str.strip() ) )     # strip(): sometimes a spurious space
                    eng_d = decimals                # using value for decimals
                    if pos_rus == 0:                # rus => decimals
                        logging.debug( "rus: %s, decimals: %d" % ( rus, decimals ) )
                        vocab[ rus ] = decimals     # using key for rus
                    else:                           # eng => decimals
                        logging.debug( "eng: %s, decimals: %d" % ( eng, decimals ) )
                        vocab[ eng ] = decimals     # using key for eng
                else:
                    # vocab is a bidict(), not dict()
                    # used for normal bidict translation
                    logging.debug( "rus: %s, eng: %s" % ( rus, eng ) )
                    rus_s = json.dumps( { "rus" : rus } )
                    eng_s = json.dumps( { "eng" : eng } )
                    vocab[ rus_s ] = eng_s
            else:
                if vocab_fname == "ERRHS_Vocabulary_histclasses.csv":
                    try:
                        byear = parts_list[ 2 ].strip()
                        dtype = parts_list[ 3 ].strip()
                    except:
                        logging.info( "csv_line: |%s|" % csv_line )
                        Nexcept += 1
                        type_, value, tb = sys.exc_info()
                        msg = "%s: %s %s" % ( type_, vocab_fname, value )
                        logging.error( msg )
                    
                    rus_d = { "rus" : rus, "byear" : byear, "dtype" : dtype }
                    eng_d = { "eng" : eng, "byear" : byear, "dtype" : dtype }
                
                elif vocab_fname == "ERRHS_Vocabulary_modclasses.csv":
                    dtype = parts_list[ 2 ][ 4: ].strip()
                    rus_d = { "rus" : rus, "dtype " : dtype  }
                    eng_d = { "eng" : eng, "dtype " : dtype  }
                else:
                    continue
                
                logging.debug( "rus: %s, eng: %s" % ( rus_d[ "rus" ], eng_d[ "eng" ] ) )
            
                rus_s = json.dumps( rus_d )
                eng_s = json.dumps( eng_d )
                
                """
                # test
                rus_d = json.loads( rus_s )
                eng_d = json.loads( eng_s )
                logging.debug( "%s rus_d: %s, %s eng_d: %s" % ( type( rus_d, ), rus_d, type( eng_d ), eng_d ) )
                logging.debug( "rus: %s, eng: %s" % ( rus_d[ "rus" ], eng_d[ "eng" ] ) )
                """
                
                try:
                    vocab[ rus_s ] = eng_s
                except:
                    Nexcept += 1
                    type_, value, tb = sys.exc_info()
                    msg = "%s: %s %s" % ( type_, vocab_fname, value )
                    logging.error( msg )
                    #sys.stderr.write( "%s\n" % msg )
                
        nline += 1
    
    vocab_file.close()

    return vocab



def convert_excel2csv( config_parser, excel_package ):
    # Convert Russian xlsx files to Russian csv files. 
    # Some filtering is applied: 
    # 1) rounding values, and 
    # 2) creating the proper ter_names from the ter_codes
    global Nexcept
    
    logging.info( "convert_excel2csv() excel_package: %s" % excel_package )
    
    tmp_dir = config_parser.get( "config", "tmppath" )
    
    xlsx_subdir = "xlsx"
    csv_subdir  = "csv-ru"

    
    # dataverse xlsx has no -ru or -en, but we want to be explicit
    extra = "-ru"   # for output csv file
    
    xlsx_basedir = os.path.join( tmp_dir, "dataverse_src", xlsx_subdir )
    
    # read regions vocab, to correct territory names via ter_code
    vocab_regions = dict()      # special, not bidict() !
    vocab_regions = load_vocab( config_parser, "ERRHS_Vocabulary_regions.csv", vocab_regions, 1, 2 )
    """
    logging.info( "ERRHS_Vocabulary_regions.csv" )
    for key, val in vocab_regions.items():
        logging.info( "key: %s, val: %s" % ( key, val ) )
    """
    
    vocab_units = dict()
    vocab_units_ru = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 0, 0 )
    #vocab_units_en = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 1, 1 )
    
    # read russian excel files
    dir_list = os.listdir( xlsx_basedir )
    dir_list.sort()
    for handle_name in dir_list:
        """
        # TEST
        if handle_name != "hdl_errhs_agriculture":
        #if handle_name != "hdl_errhs_population":
            logging.info( "skip handle_name: %s" % handle_name )
            continue
        """
        
        logging.info( "handle_name:  %s" % handle_name )
        
        xlsx_dir = os.path.join( tmp_dir, "dataverse_src", xlsx_subdir, handle_name )
        csv_dir  = os.path.join( tmp_dir, "dataverse_dst", csv_subdir,  handle_name )
       
        logging.info( "xlsx_dir: %s" % xlsx_dir )
        logging.info( "csv_dir:  %s" % csv_dir )
        
        if os.path.exists( csv_dir ):
            empty_dir( csv_dir )                # remove previous files
        if not os.path.exists( csv_dir ):
            os.makedirs( csv_dir )              # create destination dir
        
        dir_list = os.listdir( xlsx_dir )
        dir_list.sort()
        for xlsx_filename in dir_list:          # convert ru xlsx -> ru csv and save
            """
            # TEST
            if xlsx_filename != "ERRHS_4_01_data_1858.xlsx":
                logging.info( "SKIP xlsx_filename:  %s" % xlsx_filename )
                continue
            """
            
            if excel_package == "pandas":
                xlsx2csv_pandas( xlsx_dir, xlsx_filename, csv_dir, extra )
            elif excel_package == "openpyxl":
                xlsx2csv_openpyxl( xlsx_dir, xlsx_filename, csv_dir, extra, vocab_regions )
            elif excel_package == "tablib":
                #xlsx2csv_tablib( xlsx_dir, xlsx_filename, csv_dir, extra )
                xlsx2csv_tablib_filter( vocab_regions, vocab_units_ru, xlsx_dir, xlsx_filename, csv_dir, extra )
            else:
                logging.error( "excel_package: %s not supported" % excel_package )
                logging.error( "EXIT" )
                print( "EXIT" )
                sys.exit( 1 )



def xlsx2csv_pandas( xlsx_dir, xlsx_filename, csv_dir, extra ):
    logging.info( "xlsx2csv_pandas() %s" % xlsx_filename )

    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )
    if not os.path.isfile( xlsx_pathname ):
        return
    
    root, ext = os.path.splitext( xlsx_filename )
    csv_filename = root + extra + ".csv"
    csv_pathname = os.path.join( csv_dir, csv_filename )
    
    logging.debug(  "input: %s" % xlsx_pathname )
    logging.debug( "output: %s" % csv_pathname )
    
    sep = str(u'|').encode('utf-8')
    kwargs_pandas = { 
        'sep' : sep, 
        'line_terminator' : '\n' 
    }
    
    # dataverse column names
    dv_column_names = [
        "id",                   #  0
        "territory",            #  1
        "ter_code",             #  2
        "town",                 #  3
        "district",             #  4
        "year",                 #  5
        "month",                #  6
        "value",                #  7
        "value_unit",           #  8
        "value_label",          #  9
        "datatype",             # 10
        "histclass1",           # 11
        "histclass2",           # 12
        "histclass3",           # 13
        "histclass4",           # 14
        "histclass5",           # 15
        "histclass6",           # 16
        "histclass7",           # 17
        "histclass8",           # 18
        "histclass9",           # 19
        "histclass10",          # 20
        "class1",               # 21
        "class2",               # 22
        "class3",               # 23
        "class4",               # 24
        "class5",               # 25
        "class6",               # 26
        "class7",               # 27
        "class8",               # 28
        "class9",               # 29
        "class10",              # 30
        "comment_source",       # 31
        "source",               # 32
        "volume",               # 33
        "page",                 # 34
        "naborshik_id",         # 35
        "comment_naborshik",    # 36
        "base_year"             # 37
    ]
    
    df_xls = pd.read_excel( xlsx_pathname, index_col = False, convert_float = False, dtype = str )
    nrows = len( df_xls.index )
    
    nround = 0
    # spurious '.0' added to integer values that became float; 
    # re-round column PAGE (uppercase in dataverse files)
    decimals = 0
    for row in df_xls.index:
        # Check "PAGE" for all files
        old_val = df_xls.loc[ row, "PAGE" ]
        if type( old_val ) is float and not math.isnan( old_val ):
            new_val = str( long( round( old_val, decimals ) ) )
            if old_val != new_val:
                df_xls.loc[ row, "PAGE" ] = new_val
                nround += 1
        
        # HISTCLASS1 & CLASS2 integer strings for some files
        if xlsx_filename in [ "ERRHS_1_02_data_1897.xlsx", "ERRHS_1_02_data_1959.xlsx", "ERRHS_1_02_data_2002.xlsx" ]: 
            old_val = df_xls.loc[ row, "HISTCLASS1" ]
            if type( old_val ) is float and not math.isnan( old_val ):
                new_val = str( long( round( old_val, decimals ) ) )
                if old_val != new_val:
                    df_xls.loc[ row, "HISTCLASS1" ] = new_val
                    nround += 1
        
            old_val = df_xls.loc[ row, "CLASS2" ]
            if type( old_val ) is float and not math.isnan( old_val ):
                new_val = str( long( round( old_val, decimals ) ) )
                if old_val != new_val:
                    df_xls.loc[ row, "CLASS2" ] = new_val
                    nround += 1
    
    if nround == 0:
        logging.info( "nrows: %d, nothing rounded" % nrows )
    else:
        logging.info( "nrows: %d, nround: %d" % ( nrows, nround ) )
    
    df_xls.to_csv( csv_pathname, encoding = 'utf-8', index = False, **kwargs_pandas )



def xlsx2csv_openpyxl( xlsx_dir, xlsx_filename, csv_dir, extra, vocab_regions = None):
    logging.info( "xlsx2csv_openpyxl() %s" % xlsx_filename )

    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )
    if not os.path.isfile( xlsx_pathname ):
        return
    
    root, ext = os.path.splitext( xlsx_filename )
    csv_filename = root + extra + ".csv"
    csv_pathname = os.path.join( csv_dir, csv_filename )
    
    # ERRHS_1_10_data_1897.xlsx     correct losing trailing 0: 1.1 => 1.10
    # 0123456789
    datatype_file = None
    if  xlsx_filename.startswith( "ERRHS_" ) and xlsx_filename[ 9 ] == '0':
        datatype_file = "%s.%s%s"% ( xlsx_filename[ 6 ],  xlsx_filename[ 8 ],xlsx_filename[ 9 ] ) 
        logging.info( "datatype_file: %s" % datatype_file )
    
    logging.debug( "input:  %s" % xlsx_pathname )
    logging.debug( "output: %s" % csv_pathname )
    
    # ERRHS postgres column names
    pg_column_names = [
        "id",                   #  0
        "territory",            #  1
        "ter_code",             #  2
        "town",                 #  3
        "district",             #  4
        "year",                 #  5
        "month",                #  6
        "value",                #  7
        "value_unit",           #  8
        "value_label",          #  9
        "datatype",             # 10
        "histclass1",           # 11
        "histclass2",           # 12
        "histclass3",           # 13
        "histclass4",           # 14
        "histclass5",           # 15
        "histclass6",           # 16
        "histclass7",           # 17
        "histclass8",           # 18
        "histclass9",           # 19
        "histclass10",          # 20
        "class1",               # 21
        "class2",               # 22
        "class3",               # 23
        "class4",               # 24
        "class5",               # 25
        "class6",               # 26
        "class7",               # 27
        "class8",               # 28
        "class9",               # 29
        "class10",              # 30
        "comment_source",       # 31
        "source",               # 32
        "volume",               # 33
        "page",                 # 34
        "naborshik_id",         # 35
        "comment_naborshik",    # 36
        "base_year"             # 37
    ]
    ncolumns_pg = len( pg_column_names )
    
    ntername_change = 0     # count TERRITORY changes
    
    encoding  = "utf-8"
    delimiter = '|'         # notice that we imported backports.csv
    newline   = '\n'
    
    # parameter guess_types = False has been dropped (unfortunately)
    read_only = False       # need to modify some columns/cells
    workbook = openpyxl.load_workbook( xlsx_pathname, read_only = read_only, data_only = True )
    worksheet = workbook.get_active_sheet()
    
    with io.open( csv_pathname, "w", newline = newline, encoding = encoding ) as csv_file:
        writer = csv.writer( csv_file, delimiter = delimiter )
        dv_column_names = []
        
        for r, row in enumerate( worksheet.iter_rows() ):
            ter_name = None
            
            if r == 0:
                for c, cell in enumerate( row ):
                    name = cell.value
                    #logging.debug( "%d: %s" % ( c, name ) )
                    dv_column_names.append( name )
                
                cell_list = dv_column_names
                ncolumns_dv = len( dv_column_names )
                #logging.info( "number of columns: %d" % ncolumns_dv )
                if vocab_regions is not None:
                    # only check data files, not relevant for vocab files
                    if ncolumns_dv != ncolumns_pg:
                        logging.warn( "number of columns: %d, expected: %d" % ( ncolumns_dv, ncolumns_pg ) )
                
            else:
                #logging.info( "row: %d" % r )
                cell_list = []
                for c, cell in enumerate( row ):
                    column_name = dv_column_names[ c ]
                    #logging.debug( "c: %d, column: %s, value: %s" % ( c, column_name, cell.value ) )
                    
                    if vocab_regions is None:
                        # vocab files; no editing
                        cell_list.append( cell.value )
                    else:
                        # data files; edit some columns
                        if column_name == "TERRITORY":
                            ter_name = cell.value
                            # do NOT append cell.value here, together with ter_code below
                        
                        elif column_name == "TER_CODE":
                            ter_code = cell.value
                            #logging.info( "ter_code: |%s|" % ter_code )
                            terr_d = { "terr" : ter_code }
                            terr_s = json.dumps( terr_d )
                            rus_eng_s = vocab_regions.get( terr_s )
                            #logging.info( "%s: %s" % ( type( rus_eng_s ), rus_eng_s ) )
                            
                            if rus_eng_s is None:
                                logging.warn( "terr_s is not a key: %s" % terr_s )
                                rus_ter_name = ter_name     # keep existing value
                            else:
                                rus_eng = json.loads( rus_eng_s )
                                #logging.info( "%s: %s" % ( type( rus_eng ), rus_eng ) )
                                rus_ter_name = rus_eng[ u'rus' ]
                                
                                if rus_ter_name != ter_name:
                                    logging.debug( "row: %d, ter_code: %s, territory change: %s ==> %s" % ( r, ter_code, ter_name, rus_ter_name  ) )
                                    ntername_change += 1
                            
                            cell_list.append( rus_ter_name )
                            cell_list.append( ter_code )
                        
                        #elif column_name == "VALUE":
                        # round values right here ?!
                        
                        elif column_name == "DATATYPE":
                            if  datatype_file is not None:
                                cell_list.append( datatype_file )   # restore trailing '0'
                            else:
                                cell_list.append( cell.value )      # unchanged
                        else:
                            cell_list.append( cell.value )
                
            writer.writerow( cell_list )
    
    if ntername_change > 0:
        logging.info( "Number of records with change in TERRITORY column: %d" % ntername_change )



def xlsx2csv_tablib( xlsx_dir, xlsx_filename, csv_dir, extra ):
    logging.info( "xlsx2csv_tablib() %s" % xlsx_filename )

    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )
    if not os.path.isfile( xlsx_pathname ):
        return
    
    root, ext = os.path.splitext( xlsx_filename )
    csv_filename = root + extra + ".csv"
    csv_pathname = os.path.join( csv_dir, csv_filename )

    logging.debug( "input:  %s" % xlsx_pathname )
    logging.debug( "output: %s" % csv_pathname )

    encoding  = "utf-8"
    delimiter = '|'
    newline   = '\n'
    
    with io.open( csv_pathname, "w", newline = newline, encoding = encoding ) as csv_file: 
        with io.open( xlsx_pathname, "rb" ) as xlsx_file:
            data = tablib.Dataset()
            data.xlsx = xlsx_file.read()
            csv_file.write( data.export( "csv", delimiter = delimiter ) )



def xlsx2csv_tablib_filter( vocab_regions, vocab_units_ru, xlsx_dir, xlsx_filename, csv_dir, extra ):
    logging.info( "xlsx2csv_tablib() %s" % xlsx_filename )
    
    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )
    if not os.path.isfile( xlsx_pathname ):
        return
    
    root, ext = os.path.splitext( xlsx_filename )
    csv_filename = root + extra + ".csv"
    csv_pathname = os.path.join( csv_dir, csv_filename )

    logging.debug( "input:  %s" % xlsx_pathname )
    logging.debug( "output: %s" % csv_pathname )

    # ERRHS postgres column names
    pg_column_names = [
        "id",                   #  0
        "territory",            #  1
        "ter_code",             #  2
        "town",                 #  3
        "district",             #  4
        "year",                 #  5
        "month",                #  6
        "value",                #  7
        "value_unit",           #  8
        "value_label",          #  9
        "datatype",             # 10
        "histclass1",           # 11
        "histclass2",           # 12
        "histclass3",           # 13
        "histclass4",           # 14
        "histclass5",           # 15
        "histclass6",           # 16
        "histclass7",           # 17
        "histclass8",           # 18
        "histclass9",           # 19
        "histclass10",          # 20
        "class1",               # 21
        "class2",               # 22
        "class3",               # 23
        "class4",               # 24
        "class5",               # 25
        "class6",               # 26
        "class7",               # 27
        "class8",               # 28
        "class9",               # 29
        "class10",              # 30
        "comment_source",       # 31
        "source",               # 32
        "volume",               # 33
        "page",                 # 34
        "naborshik_id",         # 35
        "comment_naborshik",    # 36
        "base_year"             # 37
    ]
    ncolumns_pg = len( pg_column_names )
    
    ntername_change = 0
    nround = 0
    
    territory_loc = pg_column_names.index( "territory" )
    ter_code_loc  = pg_column_names.index( "ter_code" )
    logging.debug( "territory_loc: %d, ter_code_loc: %d" % ( territory_loc, ter_code_loc ) )
    
    value_loc    = pg_column_names.index( "value" )
    vunit_loc    = pg_column_names.index( "value_unit" )
    logging.debug( "value_loc: %d, vunit_loc: %d" % ( value_loc, vunit_loc ) )
    
    missing_units = set()
        
    encoding  = "utf-8"
    delimiter = '|'
    newline   = '\n'
    
    data_out = tablib.Dataset()
    
    with io.open( xlsx_pathname, "rb" ) as xlsx_file:
        data_in  = tablib.Dataset()
        
        data_in.xlsx = xlsx_file.read()
        
        #print( data_in.headers )
        data_out.headers = data_in.headers
        for r, row_tuple in enumerate( data_in ):       # skips header; rows are tuples
            #logging.debug( "row type: %s" % type( row_tuple ) )
            
            # row_tuple immutable; need to make changes
            row_list = list( row_tuple )
            
            # Proper ter_name from ter_code
            ter_code = row_tuple[ ter_code_loc ]
            ter_name = row_tuple[ territory_loc ]
            
            #logging.debug( "ter_code: |%s|" % ter_code )
            terr_d = { "terr" : ter_code }
            terr_s = json.dumps( terr_d )
            rus_eng_s = vocab_regions.get( terr_s )
            #logging.debug( "%s: %s" % ( type( rus_eng_s ), rus_eng_s ) )
            
            if rus_eng_s is None:
                logging.warn( "terr_s is not a key: %s" % terr_s )
                rus_ter_name = ter_name     # keep existing value
            else:
                rus_eng = json.loads( rus_eng_s )
                #logging.info( "%s: %s" % ( type( rus_eng ), rus_eng ) )
                rus_ter_name = rus_eng[ u'rus' ]
                
                if rus_ter_name != ter_name:
                    logging.debug( "row: %d, ter_code: %s, territory change: %s ==> %s" % ( r, ter_code, ter_name, rus_ter_name  ) )
                    ntername_change += 1
                    row_list[ territory_loc ] = rus_ter_name
            
            # Rounding
            vunit = row_tuple[ vunit_loc ]
            decimals = vocab_units_ru.get( vunit )
            
            if decimals is None:
                missing_units.add( vunit )
            else:
                old_val = row_tuple[ value_loc ]
                if isinstance( old_val, long ) or isinstance( old_val, float ):
                    try:
                        new_val = round( float( old_val ), decimals )
                        if decimals == 0:
                            new_val = long( new_val )
                        if old_val != new_val:
                            row_list[ value_loc ] = new_val
                            nround += 1
                            logging.debug( "row: %d, old_val: %s => new_val: %s" % ( r, old_val, new_val ) )
                    except:
                        logging.error( "xlsx2csv_tablib_filter() rounding failed" )
                        logging.info( "type: %s, old_val: |%s|" % ( type( old_val ), old_val ) )
                        logging.info( "type: %s, new_val: |%s|" % ( type( new_val ), new_val ) )
                        type_, value, tb = sys.exc_info()
                        msg = "%s: %s" % ( type_, value )
                        logging.error( msg )
                        sys.exit( 1 )
                else:
                    if isinstance( old_val, unicode ):
                        pass
                    else:
                        logging.warn( "old_val: %s, type: %s" % ( str( old_val ), type( old_val ) ) )
            
            data_out.append( row_list )
        
        nrows = len( data_in )
        
        logging.info( "nrows: %d, nterritory changes: %d, nrounded values: %d" % ( nrows, ntername_change, nround ) )
        
        nmissing = len( missing_units )
        if nmissing > 0:
            logging.warn( "missing vocab_unit keys: %d" % nmissing )
            for unit in missing_units:
                logging.warn( unit )
    
    with io.open( csv_pathname, "w", newline = newline, encoding = encoding ) as csv_file:
        csv_file.write( data_out.export( "csv", delimiter = delimiter ) )



def translate_errhs_csvs( config_parser, handle_names ):
    logging.info( "translate_errhs_csvs()" )
    vocab_units = bidict()
    vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 0, 1 )  # ru => en
    #vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 0, 1 )  # en => ru
    
    vocab_regions = dict()      # special, not bidict() !
    vocab_regions = load_vocab( config_parser, "ERRHS_Vocabulary_regions.csv", vocab_regions, 1, 2 )
    
    vocab_histclasses = bidict()
    vocab_fname = "ERRHS_Vocabulary_histclasses.csv"
    vocab_histclasses = load_vocab( config_parser, vocab_fname, vocab_histclasses, 0, 1 )
    logging.info( "check vocab_fname: %s" % vocab_fname )
    for key in vocab_histclasses:
        logging.debug( "key: %s, value: %s" % ( key, vocab_histclasses[ key ] ) )
    
    vocab_modclasses = bidict()
    vocab_modclasses = load_vocab( config_parser, "ERRHS_Vocabulary_modclasses.csv", vocab_modclasses, 0, 1 )
    
    # TODO process existing files, not trying all handle_names
    # csv-ru/hdl_errhs_[type]/ERRHS_[datatype]_data_[year]-ru.csv2xlsx
    # csv-en/hdl_errhs_[type]/ERRHS_[datatype]_data_[year]-en.csv
    for handle_name in handle_names:
        translate_csv( config_parser, handle_name, vocab_units, vocab_regions, vocab_histclasses, vocab_modclasses )



def translate_csv( config_parser, handle_name, vocab_units, vocab_regions, vocab_histclasses, vocab_modclasses ):
    global Nexcept
    
    logging.info( "translate_csv()" )
    logging.info( "translating csv documents for handle name %s ..." % handle_name )
    
    tmp_dir = config_parser.get( "config", "tmppath" )
    csv_dir = os.path.join( tmp_dir, "dataverse_dst", "csv-ru", handle_name )
    if os.path.exists( csv_dir ):
        logging.debug( "csv_dir: %s" % csv_dir )
    else:
        logging.debug( "not found, skip: csv_dir: %s" % csv_dir )
        return
    
    eng_dir = os.path.join( tmp_dir, "dataverse_dst", "csv-en", handle_name )
    logging.debug( "eng_dir: %s" % eng_dir )
    
    if os.path.exists( eng_dir ):
        empty_dir( eng_dir )                # remove previous files
    if not os.path.exists( eng_dir ):
        os.makedirs( eng_dir )
    
    regions     = [ "TERRITORY" ]
    units       = [ "VALUE_UNIT" ]
    histclasses = [ "HISTCLASS1", "HISTCLASS2", "HISTCLASS3", "HISTCLASS4", "HISTCLASS5",  "HISTCLASS6", "HISTCLASS7", "HISTCLASS8", "HISTCLASS9", "HISTCLASS10", ]
    modclasses  = [ "CLASS1", "CLASS2", "CLASS3", "CLASS4", "CLASS5",  "CLASS6", "CLASS7", "CLASS8", "CLASS9", "CLASS10", ]
    translate_list = regions + units + histclasses + modclasses
    
    # read russian csv files
    dir_list = os.listdir( csv_dir )
    dir_list.sort()
    for csv_name in dir_list:
        #if csv_name != "ERRHS_4_01_data_1858-ru.csv":
        #    logging.info( "SKIP: %s" % csv_name )
        #    continue
        
        nexceptions = 0
        exceptions = []
        
        #logging.info( csv_name )
        csv_path = os.path.join( csv_dir, csv_name )
        if not os.path.isfile( csv_path ):
            continue
        
        logging.debug( csv_path )
        basename, ext = os.path.splitext( csv_name )
        if basename.endswith( "-ru" ):
            basename = basename[ :-3 ]
        eng_name = basename + "-en" +  ext
        logging.info( "%s => %s" % ( csv_name, eng_name ) )
        
        eng_path = os.path.join( eng_dir, eng_name )
        logging.debug( eng_path )
        
        file_rus = open( csv_path, "r" )
        file_eng = open( eng_path, "w" )
        #file_rus = codecs.open( csv_path, "r", encoding = 'utf-8' )
        #file_eng = codecs.open( eng_path, "w", encoding = 'utf-8' )
        
        not_found = set()
        nline = 0
        for csv_line in iter( file_rus ):
            csv_line = csv_line.strip()
            logging.debug( "csv_line: %s" % csv_line )
            eng_cols = []
            if nline == 0:
                header = csv_line.split( '|' )
                logging.debug( header )
                file_eng.write( "%s\n" % csv_line )     # header is english, only data fields must be translated
            else:
                rus_cols = csv_line.split( '|' )
                ncolumns = len( rus_cols )
                for c in range( ncolumns ):
                    name = header[ c ]
                    rus_str = rus_cols[ c ]
                    
                    vocab = None
                    vocab_name = None
                    if name in units:
                        vocab = vocab_units
                        vocab_name = "vocab_units"
                        rus_d = { "rus" : rus_str }
                    
                    elif name in regions:
                        vocab = vocab_regions
                        vocab_name = "vocab_regions"
                        i_terr = header.index( "TER_CODE" )
                        terr = rus_cols[ i_terr ]
                        terr_d = { "terr" : terr }
                    
                    elif name in histclasses:
                        vocab = vocab_histclasses
                        vocab_name = "histclasses"
                        i_byear = header.index( "BASE_YEAR" )
                        i_dtype = header.index( "DATATYPE" )
                        
                        """
                        len_header = len( header )
                        #if i_byear + 1 > len_header:
                        logging.debug( "header length: %d" % len_header )
                        logging.debug( "i_byear: %d" % i_byear )
                        logging.debug( "rus_cols length: %d" % len( rus_cols ) )
                        """
                        
                        byear = rus_cols[ i_byear ]
                        dtype = rus_cols[ i_dtype ]
                        rus_d = { "rus" : rus_str, "byear" : byear, "dtype" : dtype }
                    
                    elif name in modclasses:
                        vocab_name = "modclasses"
                        vocab = vocab_modclasses
                        dtype = rus_cols[ i_dtype ]
                        rus_d = { "rus" : rus_str, "dtype " : dtype  }
                    
                    if rus_str in [ "", ".", ". " ]:
                        eng_str = rus_str
                    #elif rus_str in [ "test1", "test2", "test3", "test4", "test5", "test6", "test7", "test8", "test9", "test10" ]:
                    #    eng_str = rus_str
                    elif rus_str.isdigit():     # only digits
                        eng_str = rus_str
                    elif vocab is not None:
                        if name in regions:
                            terr_s = json.dumps( terr_d )
                            logging.debug( "translate: vocab_name: %s, %s %s, key: %s" % ( vocab_name, name, rus_str, terr_s ) )
                            try:
                                rus_eng_val = vocab[ terr_s ]
                                logging.debug( "rus_eng_val: %s" % rus_eng_val )
                                rus_eng_d = json.loads( rus_eng_val )
                                eng_str = rus_eng_d[ "eng" ]
                                logging.debug( "translate: %s %s => %s" % ( name, rus_str, eng_str ) )
                            except:
                                Nexcept += 1
                                eng_str = rus_str
                                nexceptions += 1
                                type_, value, tb = sys.exc_info()
                                msg = "%s: %s" % ( type_, value )
                                logging.error( msg )
                                #sys.stderr.write( "%s\n" % msg )
                        else:
                            rus_key = json.dumps( rus_d )
                            logging.debug( "translate: vocab_name: %s, %s %s, key: %s" % ( vocab_name, name, rus_str, rus_key ) )
                            try:
                                eng_val = vocab[ rus_key ]
                                eng_d = json.loads( eng_val )
                                eng_str = eng_d[ "eng" ]
                                logging.debug( "translate: %s %s => %s" % ( name, rus_str, eng_str ) )
                            except:
                                Nexcept += 1
                                eng_str = rus_str
                                nexceptions += 1
                                type_, value, tb = sys.exc_info()
                                msg = "%s: (%s) %s" % ( type_, rus_d[ "rus" ], value )
                                if msg not in exceptions:        # only disply new exceptions
                                    logging.error( msg )
                                    #sys.stderr.write( "%s\n" % msg )
                                    exceptions.append( msg )
                    else:
                        eng_str = rus_str
                    
                    if eng_str is None:
                        eng_str = rus_str
                        msg = "Not found in %s: translation for rus_str: %s" % ( csv_name, rus_str )
                        #if msg not in not_found:
                        not_found.add( msg )
                    
                    eng_cols.append( eng_str )
                
                eng_line = '|'.join( eng_cols )
                logging.debug( eng_line )
                file_eng.write( "%s\n" % eng_line )
                
            nline += 1
        
        if nexceptions != 0:
            logging.error( "translate_csv: number of exceptions: %d" % nexceptions )
        if len( not_found ) > 0:
            logging.error( "translate_csv: unique not_found strings: %d" % len( not_found ) )
        
        #not_found = list( set( not_found ) )
        for item in not_found:
            msg = "not found: %s" % item
            logging.info( msg )
            #sys.stderr.write( "%s\n" % msg )
        
        file_rus.close()
        file_eng.close()



def compile_filecatalogue( config_parser, language, excel_package, pd_engine ):
    # convert the -en and -ru csv files to xlsx files with copyright tab
    global Nexcept
    
    logging.info( "compile_filecatalogue() language = %s" % language )
    
    # rounding of data in value column is specified in ERRHS_Vocabulary_units.csv
    # Equality of pos_rus and pos_eng is a hack, used to flag decimals from either rus or eng
    config_parser = get_configparser()
    
    vocab_units = dict()
    if language == "ru":
        vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 0, 0 )
    elif language == "en":
        vocab_units = load_vocab( config_parser, "ERRHS_Vocabulary_units.csv", vocab_units, 1, 1 )
    else:
        logging.error( "language must be either 'ru' or 'en'" )
    
    #logging.info( "vocab_units: %s" % str( vocab_units ) )
    """
    logging.info( "vocab_units keys:" )
    for key in vocab_units:
        logging.info( key )
    """
    
    tmp_dir = config_parser.get( "config", "tmppath" )
    
    csv_subdir  = "csv-"  + language
    fcat_subdir = "fcat-" + language
    
    csv_basedir = os.path.join( tmp_dir, "dataverse_dst", csv_subdir )
    
    dir_list = os.listdir( csv_basedir )
    dir_list.sort()
    for handle_name in dir_list:
        """
        #if handle_name != "hdl_errhs_land":     # en OK
        if handle_name != "hdl_errhs_labour":
            logging.info( "skip handle_name:  %s" % handle_name )
            continue
        """
        
        logging.info( "handle_name:  %s" % handle_name )
        
        csv_dir  = os.path.join( tmp_dir, "dataverse_dst", csv_subdir, handle_name )
        xlsx_dir = os.path.join( tmp_dir, "dataverse_dst", fcat_subdir, handle_name )   # _dst because not from dataverse retrieval
        
        logging.info( "csv_dir:  %s" % csv_dir )
        logging.info( "xlsx_dir: %s" % xlsx_dir )
        
        if os.path.exists( xlsx_dir ):
            empty_dir( xlsx_dir )                # remove previous files
        if not os.path.exists( xlsx_dir ):
            os.makedirs( xlsx_dir )              # create destination dir
        
        dir_list = os.listdir( csv_dir )
        dir_list.sort()
        for csv_filename in dir_list:
            # convert csv -> xlsx and save
            
            if excel_package == "pandas":
                csv2xlsx_pandas( language, vocab_units, csv_dir, csv_filename, xlsx_dir, pd_engine )
            elif excel_package == "tablib":
                csv2xlsx_tablib( language, vocab_units, csv_dir, csv_filename, xlsx_dir )
            else:
                logging.error( "excel_package: %s not supported" % excel_package )
                logging.error( "EXIT" )
                print( "EXIT" )
                sys.exit( 1 )



def csv2xlsx_pandas( language, vocab_units, csv_dir, csv_filename, xlsx_dir, pd_engine ): 
    logging.info( "csv2xlsx_pandas() %s" % csv_filename )
    
    """
    autoupdate.py:1703: UnicodeWarning: Unicode equal comparison failed to convert both arguments to Unicode - interpreting them as being unequal
    if old_val == '.':
    sys:1: DtypeWarning: Columns (7) have mixed types. Specify dtype option on import or set low_memory=False.
    sys:1: DtypeWarning: Columns (7,34) have mixed types. Specify dtype option on import or set low_memory=False.
    sys:1: DtypeWarning: Columns (9) have mixed types. Specify dtype option on import or set low_memory=False.
    sys:1: DtypeWarning: Columns (11,22) have mixed types. Specify dtype option on import or set low_memory=False.
    """
    
    #if csv_filename != "ERRHS_2_01_data_1795-en.csv":
    #    logging.info( "skip input: %s" % csv_filename )
    #    return
    
    # ERRHS_1_10_data_1897.xlsx     correct losing trailing 0: 1.1 => 1.10
    # 0123456789
    datatype_file = None
    if csv_filename[ 9 ] == '0' and csv_filename.startswith( "ERRHS_" ):
        datatype_file = "%s.%s%s"% ( csv_filename[ 6 ],  csv_filename[ 8 ],csv_filename[ 9 ] ) 
        logging.info( "datatype_file: %s" % datatype_file )
    
    csv_pathname = os.path.join( csv_dir, csv_filename )
    if not os.path.isfile( csv_pathname ):
        return
    
    logging.debug( "input: %s" % csv_pathname )
    
    # dataverse column names
    dv_column_names = [
        "id",                   #  0
        "territory",            #  1
        "ter_code",             #  2
        "town",                 #  3
        "district",             #  4
        "year",                 #  5
        "month",                #  6
        "value",                #  7
        "value_unit",           #  8
        "value_label",          #  9
        "datatype",             # 10
        "histclass1",           # 11
        "histclass2",           # 12
        "histclass3",           # 13
        "histclass4",           # 14
        "histclass5",           # 15
        "histclass6",           # 16
        "histclass7",           # 17
        "histclass8",           # 18
        "histclass9",           # 19
        "histclass10",          # 20
        "class1",               # 21
        "class2",               # 22
        "class3",               # 23
        "class4",               # 24
        "class5",               # 25
        "class6",               # 26
        "class7",               # 27
        "class8",               # 28
        "class9",               # 29
        "class10",              # 30
        "comment_source",       # 31
        "source",               # 32
        "volume",               # 33
        "page",                 # 34
        "naborshik_id",         # 35
        "comment_naborshik",    # 36
        "base_year"             # 37
    ]
    
    
    #delimiter = b'|'       # leading b required with __future__, otherwise: TypeError: "delimiter" must be an 1-character string
    #delimiter = str( '|' )  # TypeError: "delimiter" must be string, not newstr
    delimiter = str( u'|' ).encode( 'utf-8' )
    
    """
    with open( csv_pathname, "rb" ) as csv_file:
    #with codecs.open( csv_pathname, "rb", encoding = 'utf-8' ) as csv_file:
        csv_reader = csv.reader( csv_file, delimiter = delimiter )
        for row in csv_reader:
            #logging.debug( ", ".join( row ) )
            logging.debug( str( row ) )
    """
    
    #sep = str( u'|' ).encode( "utf-8" )            # "encode method has been disabled in newbytes"
    #sep = str( delimiter ).encode( "utf-8" )       # "encode method has been disabled in newbytes"
    sep = b'|'
    
    kwargs_pandas = \
    {
        "encoding" : "utf-8",
        "sep"      : sep            # no effect
        #,"dtype"    : { "page" : "object" }
        ,"dtype"    : "str"         # how to prevent stupid int to float coversion?
        #,"line_terminator" : '\n'   # TypeError: parser_f() got an unexpected keyword argument "line_terminator"
    }
    
    # read csv into pandas dataframe 1
    df1 = pd.read_csv( csv_pathname, **kwargs_pandas )
    nrows = len( df1.index )
    
    missing_units = set()
    
    nround = 0
    # spurious '.0' added to integer values; re-round column VALUE (uppercase in csv)
    for row in df1.index:
        unit = df1.loc[ row, "VALUE_UNIT" ]
        decimals = vocab_units.get( unit )
        if decimals is None:
            missing_units.add( unit )
            continue
        elif decimals == 0:
            old_val = df1.loc[ row, "VALUE" ]
            #if old_val == '.': # UnicodeWarning: Unicode equal comparison failed to convert both arguments to Unicode - interpreting them as being unequal
            if old_val.strip() == u'.':
                continue
            
            try:
                new_val = str( long( round( float( old_val ), decimals ) ) )
                if old_val != new_val:
                    df1.loc[ row, "VALUE" ] = new_val
                    nround += 1
            except:
                logging.error( "csv2xlsx() rounding failed" )
                logging.info( "type: %s, old_val: |%s|" % ( type( old_val ), old_val ) )
                logging.info( "type: %s, new_val: |%s|" % ( type( new_val ), new_val ) )
                type_, value, tb = sys.exc_info()
                msg = "%s: %s" % ( type_, value )
                logging.error( msg )
        else:
            continue
    
        #page = df1.loc[ row, "PAGE" ]
        
        if datatype_file is not None:       # 1.1 => 1.10 restore
            df1.loc[ row, "DATATYPE" ] = datatype_file
    
    
    if nround == 0:
        logging.info( "nrows: %d, nothing rounded" % nrows )
    else:
        logging.info( "nrows: %d, nround: %d" % ( nrows, nround ) )
    
    nmissing = len( missing_units )
    if nmissing > 0:
        logging.warn( "missing vocab_unit keys: %d" % nmissing )
        for unit in missing_units:
            logging.warn( unit )
    
    """
    # sort by ter_code and histclasses
    sort_columns = []
    sort_columns.append( "TER_CODE" )
    for l in range( 10 ):
        l_str = "HISTCLASS%d" % ( l + 1 )
        sort_columns.append( l_str )
    
    logging.debug( "sort by: %s" % sort_columns )
    df1 = df1.sort_values( by = sort_columns, ascending = False )
    """
    
    root, ext = os.path.splitext( csv_filename )
    xlsx_filename = root + ".xlsx"
    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )
    
    # from input parameter
    #pd_engine = "openpyxl"
    #pd_engine = "xlsxwriter"
    
    writer = pd.ExcelWriter( xlsx_pathname, engine = pd_engine )
    
    try:
        df1.to_excel( writer, "Table", encoding = "utf-8", index = False )
    except:
        logging.error( "convert_csv2xlsx() to_excel failed: %s" % csv_filename )
        type_, value, tb = sys.exc_info()
        msg = "%s: %s %s" % ( type_, xlsx_filename, value )
        logging.error( msg )
        
    # create copyright sheet in pandas dataframe 2
    if language == "en":
        df2 = pd.DataFrame( { " ": [ 
            "",
            "",
            "Creative Commons License", 
            "This work is licensed under a Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License.", 
            "http://creativecommons.org/licenses/by-nc-sa/4.0/", 
            "", 
            "By downloading and using data from the Electronic Repository of Russian Historical Statistics the user agrees to the terms of this license. Providing a correct reference to the resource is a formal requirement of the license: ", 
            "Kessler, Gijs and Andrei Markevich (%d), Electronic Repository of Russian Historical Statistics, 18th - 21st centuries, http://ristat.org/" % date.today().year, 
        ] } )
        """
        c = ws_cr.cell( row = 4, column = 0 )
        c.value = "Creative Commons License"
        c = ws_cr.cell( row = 5, column = 0 )
        c.value = "This work is licensed under a Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License."
        c = ws_cr.cell( row = 6, column = 0 )
        c.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/"
        c = ws_cr.cell( row = 8, column = 0 )
        c.value = "By downloading and using data from the Electronic Repository of Russian Historical Statistics the user agrees to the terms of this license. Providing a correct reference to the resource is a formal requirement of the license: "
        c = ws_cr.cell( row = 9, column = 0 )
        c.value = "Kessler, Gijs and Andrei Markevich (%d), Electronic Repository of Russian Historical Statistics, 18th - 21st centuries, http://ristat.org/" % date.today().year
        """
    elif language == "ru":
        df2 = pd.DataFrame( { " ": [ 
            "",
            "",
            "Лицензия Creative Commons", 
            "Это произведение доступно по лицензии Creative Commons «Attribution-NonCommercial-ShareAlike» («Атрибуция — Некоммерческое использование — На тех же условиях») 4.0 Всемирная.", 
            "http://creativecommons.org/licenses/by-nc-sa/4.0/deed.ru", 
            "",
            "Скачивая и начиная использовать данные пользователь автоматически соглашается с этой лицензией. Наличие корректно оформленной ссылки является обязательным требованием лицензии:", 
            "Кесслер Хайс и Маркевич Андрей (%d), Электронный архив Российской исторической статистики, XVIII – XXI вв., [Электронный ресурс] : [сайт]. — Режим доступа: http://ristat.org/" % date.today().year, 
        ] } )
        """
        c = ws_cr.cell( row = 4, column = 0 )
        c.value = "Лицензия Creative Commons"
        c = ws_cr.cell( row = 5, column = 0 )
        c.value = "Это произведение доступно по лицензии Creative Commons «Attribution-NonCommercial-ShareAlike» («Атрибуция — Некоммерческое использование — На тех же условиях») 4.0 Всемирная."
        c = ws_cr.cell( row = 6, column = 0 )
        c.value = "http://creativecommons.org/licenses/by-nc-sa/4.0/deed.ru"
        c = ws_cr.cell( row = 8, column = 0 )
        c.value = "Скачивая и начиная использовать данные пользователь автоматически соглашается с этой лицензией. Наличие корректно оформленной ссылки является обязательным требованием лицензии:"
        c = ws_cr.cell( row = 9, column = 0 )
        c.value = "Кесслер Хайс и Маркевич Андрей (%d), Электронный архив Российской исторической статистики, XVIII – XXI вв., [Электронный ресурс] : [сайт]. — Режим доступа: http://ristat.org/" % date.today().year
        """
    
    # Convert the dataframe to an XlsxWriter Excel object.
    logging.debug( "output: %s" % xlsx_filename )
    df2.to_excel( writer, sheet_name = "Copyrights", encoding = "utf-8", index = False )
    
    writer.save()
    writer.close()



def csv2xlsx_tablib( language, vocab_units, csv_dir, csv_filename, xlsx_dir ):
    logging.info( "csv2xlsx_tablib() %s" % csv_filename )

    csv_pathname = os.path.join( csv_dir, csv_filename )
    if not os.path.isfile( csv_pathname ):
        return
    
    root, ext = os.path.splitext( csv_filename )
    xlsx_filename = root + ".xlsx"
    xlsx_pathname = os.path.join( xlsx_dir, xlsx_filename )

    logging.debug( "input:  %s" % csv_pathname )
    logging.debug( "output: %s" % xlsx_pathname )

    encoding  = "utf-8"     # not a parameter of data.load()
    delimiter = '|'
    newline   = '\n'        # not a parameter of data.load()
    
    with io.open( xlsx_pathname, "wb" ) as xlsx_file:
        with io.open( csv_pathname, "r" ) as csv_file:
            data = tablib.Dataset()
            try:
                data.load( csv_file.read(), format = "csv", delimiter = delimiter )
                xlsx_file.write( data.export( "xlsx" ) )
            except:
                logging.error( "csv2xlsx_tablib() load failed: %s" % csv_filename )
                type_, value, tb = sys.exc_info()
                msg = "%s: %s %s" % ( type_, xlsx_filename, value )
                logging.error( msg )
                logging.warn( csv_file.readline() )
                sys.exit( 1 )



def format_secs( seconds ):
    nmin, nsec  = divmod( seconds, 60 )
    nhour, nmin = divmod( nmin, 60 )

    if nhour > 0:
        str_elapsed = "%d:%02d:%02d (hh:mm:ss)" % ( nhour, nmin, nsec )
    else:
        if nmin > 0:
            str_elapsed = "%02d:%02d (mm:ss)" % ( nmin, nsec )
        else:
            str_elapsed = "%d (sec)" % nsec

    return str_elapsed



if __name__ == "__main__":
    DO_RETRIEVE_VOCAB    = True   # -01- vocabulary: dataverse => local_disk
    DO_RETRIEVE_DOC      = True   # -02- documentation: dataverse  => local_disk
    DO_RETRIEVE_ERRHS    = True   # -03- ERRHS data: dataverse => local_disk
    
    DO_DOC_COPY          = True   # -04- local_disk doc srx dir => doc dst dir
    DO_CONVERT_VOCAB2CSV = True   # -05- convert vocab xlsx files [ru + en] to vocab csv files [ru + en]
    DO_CONVERT_EXCEL2CSV = True   # -06- convert Russian ERRHS xlsx files to Russian ERRHS csv files
    DO_TRANSLATE_CSV     = True   # -07- translate Russian csv files to English csv files
    
    DO_POSTGRES_DB       = True   # -08- ERRHS data: local_disk => postgresql, csv -> table
    DO_MONGO_DB          = True   # -09- ERRHS data: postgresql => mongodb
    DO_FILE_CATALOGUE    = True   # -10- ERRHS data: csv -> filecatalogue xlsx
    
    #dv_format = ""
    dv_format = "original"  # does not work for ter_code (regions) vocab translations
    
    log_file = True
    
    #log_level = logging.DEBUG
    log_level = logging.INFO
    #log_level = logging.WARNING
    #log_level = logging.ERROR
    #log_level = logging.CRITICAL
    
    log_format = "%(asctime)s %(levelname)-8s %(message)s"
    log_datefmt = "%Y-%m-%d %H:%M:%S"
    
    if log_file:
        mode = 'w'
        #mode = 'a'      # debugging
        logging_filename = "autoupdate.log"
        logging.basicConfig( filename = logging_filename, filemode = mode, level = log_level, format = log_format, datefmt = log_datefmt )
    else:
        logging.basicConfig( level = log_level )

    time0 = time()      # seconds since the epoch
    logging.info( "start: %s" % datetime.now() )
    logging.info( "user: %s" % getpass.getuser() )
    logging.info( __file__ )
    
    python_vertuple = sys.version_info
    python_version = str( python_vertuple[ 0 ] ) + '.' + str( python_vertuple[ 1 ] ) + '.' + str( python_vertuple[ 2 ] )
    logging.info( "Python version: %s" % python_version )
    
    AUTOUPDATE = os.environ[ "AUTOUPDATE" ]
    try:
        AUTOUPDATE = int( AUTOUPDATE )
    except:
        AUTOUPDATE = 0
    
    logging.info( "AUTOUPDATE: %d" % AUTOUPDATE )
    if AUTOUPDATE == 0:
        logging.error( "Skipping autoupdate" )
        logging.error( "EXIT" )
        sys.exit( 0 )
    
    RUSSIANREPO_CONFIG_PATH = os.environ[ "RUSSIANREPO_CONFIG_PATH" ]
    logging.info( "RUSSIANREPO_CONFIG_PATH: %s" % RUSSIANREPO_CONFIG_PATH )
    if not os.path.isfile( RUSSIANREPO_CONFIG_PATH ):
        logging.error( "Required config file does not exist" )
        logging.error( "EXIT" )
        sys.exit( 1 )
    
    config_parser = configparser.RawConfigParser()
    config_parser.read( RUSSIANREPO_CONFIG_PATH )
    mongo_client  = MongoClient()
    
    
    # DATAVERSE RETRIEVAL Phase; will exit() on retrieval error
    if AUTOUPDATE == 1:
        autoupdate = check_autoupdate( config_parser, dv_format )
    
        if autoupdate is False:
            logging.info( "__main__: autoupdate cancelled!" )
            
            # This should be called at application exit,
            # and no further use of the logging system should be made after this call.
            logging.shutdown()
            sys.exit( 0 )
    
    if DO_RETRIEVE_VOCAB:
        time0_ = time()
        logging.info( '' )
        logging.info( "-1- DO_RETRIEVE_VOCAB" )
        # Downloaded vocabulary documents are not first stored in postgreSQL, 
        # they are processed on the fly, and directly put in MongoDB vocabulary
        
        retrieve_vocabularies( config_parser, dv_format )
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "VOCAB processing took %s" % str_elapsed )
    
    
    if DO_RETRIEVE_DOC:
        time0_ = time()
        logging.info( '' )
        logging.info( "-2- DO_RETRIEVE_DOC" )
        update_documentation( config_parser )
    
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "DOC processing took %s" % str_elapsed )
    
    if DO_RETRIEVE_ERRHS:
        time0_ = time()
        logging.info( '' )
        logging.info( "-3- DO_RETRIEVE_ERRHS" )
        for handle_name in handle_names:
            retrieve_handle_docs( config_parser, handle_name, dv_format ) # dataverse  => local_disk
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "ERRHS processing took %s" % str_elapsed )
        
    # PROCESSING Phase
    if DO_DOC_COPY: 
        time0_ = time()
        logging.info( '' )
        logging.info( "-4- DO_RETRIEVE_COPY" )
        copy_doc_src2dst()
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "DOC_COPY processing took %s" % str_elapsed )
    
    if DO_CONVERT_VOCAB2CSV:    # convert vocab xlsx files [ru + en] to vocab csv files [ru + en]
        time0_ = time()
        logging.info( '' )
        logging.info( "-5- DO_CONVERT_VOCAB2CSV" )
        
        #excel_package = "pandas"        # years get '.0'
        #excel_package = "xlrd"          # not implemented
        #excel_package = "openpyxl"      # applies losing trailing 0's correction
        excel_package = "tablib"        # from Kenneth Reitz (requests)
        
        convert_vocabularies2csv( excel_package )
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "VOCAB2CSV processing took %s" % str_elapsed )
    
    if DO_CONVERT_EXCEL2CSV:    # convert Russian ERRHS xlsx files to Russian ERRHS csv files
        time0_ = time()
        logging.info( '' )
        logging.info( "-6- DO_CONVERT_EXCEL2CSV" )
        
        #excel_package = "pandas"        # too slow (due to post-read coorections)
        #excel_package = "xlrd"          # not implemented
        #excel_package = "openpyxl"      # applies losing trailing 0's correction
        excel_package = "tablib"        # from Kenneth Reitz (requests)
        
        convert_excel2csv( config_parser, excel_package )
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "EXCEL2CSV processing took %s" % str_elapsed )
    
    if DO_TRANSLATE_CSV:
        time0_ = time()
        logging.info( '' )
        logging.info( "-7- DO_TRANSLATE_CSV" )                      # ru => en
        translate_errhs_csvs( config_parser, handle_names )
    
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "TRANSLATE_CSV processing took %s" % str_elapsed )
    
    if DO_POSTGRES_DB:
        # TODO do retrieve again
        # TODO rewrite filter_csv
        time0_ = time()
        logging.info( '' )
        logging.info( "-8- DO_POSTGRES_DB" )    # read russian csv files
        logging.StreamHandler().flush()
        for language in [ "ru", "en" ]:
            row_count( config_parser, language )
            clear_postgres_table( config_parser, language )
            row_count( config_parser, language )
            
            for handle_name in handle_names:
                store_handle_docs( config_parser, handle_name, language )   # local_disk => postgresql
                logging.StreamHandler().flush()
                row_count( config_parser,language )
            
            # done on-the-fly in services/topic_counts()
            #topic_counts( config_parser )                                  # postgresql datasets.topics counts
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "POSTGRES_DB processing took %s" % str_elapsed )
    
    if DO_MONGO_DB:
        time0_ = time()
        logging.info( '' )
        logging.info( "-9- DO_MONGO_DB" )
        mongo_store_vocabularies()
        
        for language in [ "ru", "en" ]:
            update_handle_docs( config_parser, mongo_client, language )     # postgresql => mongodb
        
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "MONGO_DB processing took %s" % str_elapsed )
    
    if DO_FILE_CATALOGUE:
        time0_ = time()
        logging.info( '' )
        logging.info( "-10- DO_FILE_CATALOGUE" )
        
        #excel_package = "pandas"        # too slow (due to post-read coorections)
        excel_package = "tablib"        # from Kenneth Reitz (requests)
        
        #pd_engine = "openpyxl"      # only for pandas
        pd_engine = "xlsxwriter"    # only for pandas
    
        for language in [ "ru" ]:  # test
        #for language in [ "en" ]:  # test
        #for language in [ "ru", "en" ]:DELETE FROM links_original.registration_o WHERE id_source = 250 AND registration_maintype = 1;

            compile_filecatalogue( config_parser, language, excel_package, pd_engine )  # create filecatalogue xlsx files
    
        str_elapsed = format_secs( time() - time0_ )
        logging.info( "FILE_CATALOGUE processing took %s" % str_elapsed )
    
    
    logging.info( '' )
    logging.info( "total number of exceptions: %d" % Nexcept )
    
    logging.info( "stop: %s" % datetime.now() )
    str_elapsed = format_secs( time() - time0 )
    logging.info( "processing took %s" % str_elapsed )
    
    # This should be called at application exit,
    # and no further use of the logging system should be made after this call.
    logging.shutdown()
    
# [eof]
